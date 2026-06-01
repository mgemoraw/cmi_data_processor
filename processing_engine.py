

from operator import index
import os
from pathlib import Path
import re
import shutil
from copy import copy 
from openpyxl import Workbook, load_workbook
from datetime import datetime 
from mappings import COLUMN_MAPPINGS

class DataProcessingEngine:
    def __init__(self, input_folder=None, template_path=None, logger=print, equipment=None):
        self.source_path = Path(input_folder) if input_folder else Path.cwd()
        self.template_path = Path(template_path) if template_path else Path.cwd()
        self.template_mapping = {}

        self.logger = logger
        self.equipment = equipment 
        self.progress_callback = None
        
        self.output_folder = os.path.join(self.source_path, "output")

        # Pattern for DD-MM-YYYY
        # self.date_pattern = r"(\d{2}-\d{2}-\d{4})"
        self.date_pattern = r"(\d{2}[-_/]\d{2}[-_/]\d{4})"
        self.data_count_pattern  = r"(\d{2}[-_/]\d{2}[-_/]\d{2}[-_/]\d{4})"

        self.SHEET_ALIASES = {
            "daily_variables": [
                "daily_variables",
                "daily variable",
                "daily variables",
                "daily-variables",
            ],

            "mpdm": [
                "mpdm",
                "MPDM"
            ]
        }

    def log(self, message):
        self.logger(message)


    def update_progress(self, value):
        if self.progress_callback:
            self.progress_callback(value)

    def get_count_and_date_from_filename(self, file_name):
        match = re.search(self.data_count_pattern, file_name)
        if match:
            count = match.group(1)[:2] # Extract the first two digits as data count
            if not match.group(2):
                date = match.group(1)[-10:]  # Extract the date part (last 10 characters)return count, date
            else:
                date = match.group(2)

            return count, date
        return None, None

    def sort_files_by_date(self, files):
        def extract_date_object(file_name):
            match = re.search(self.date_pattern, file_name)

            if match:
                date_str = match.group(1)  # Extracts the date substring (e.g., "02-03-2014")
                try:
                    # Convert the string to a real datetime object for accurate chronological sorting
                    # Adjust "%d-%m-%Y" if your pattern captures dates differently (e.g., "%Y-%m-%d")
                    return datetime.strptime(date_str, "%d-%m-%Y")
                except ValueError:
                    # Handle cases where the regex matches a string that isn't a valid calendar date
                    return datetime.max
            
            # If no date pattern is found, return datetime.max to safely push the file to the end of the list
            return datetime.max

        # Sort using the actual datetime object instead of a string
        return sorted(files, key=extract_date_object)
    
    def get_equipment_config(self):

        equipment = self.equipment.lower().strip()

        if not equipment:
            raise ValueError("Equipment type is empty")

        if equipment not in COLUMN_MAPPINGS:
            raise ValueError(
                f"Unsupported equipment type: {equipment}"
            )

        return equipment, COLUMN_MAPPINGS[equipment]
    
    def list_excel_files(self):
        """
        Return all Excel files in the folder
        """
        return list(self.source_path.glob("*.xlsx"))
    
    def read_excel_contents(self):
        self.logger("🚀 Starting data processing...")
        # Implement the main processing logic here
        self.logger(f"📂 Source folder: {self.source_path}")
        # template = self.read_template_file(self.template_path)
        files = self.sort_files_by_date(os.listdir(self.source_path))
        total_files = self.list_excel_files()
        
        for index, file in enumerate(files):
            if file.endswith('.xlsx') and file != self.template_path.name:
                self.logger(f"📄 Processing file: {file}")
                # Implement your data processing logic here
                # For example, you can read the file, extract data, etc.
                # You can also update progress using self.update_progress(value)

                count, date_str = self.get_count_and_date_from_filename(file)
                if date_str:
                    self.logger(f"📅 Extracted date from filename: {date_str}")
                
                if count:
                    self.logger(f"🔢 Extracted data count from filename: {count}")
                    data_count = count
                    


                if date_str is None :
                    date_str = self.get_date_str(os.path.join(self.source_path, file))
                
                if count is None:
                    data_count = self.format_data_count(data_count=index+1)
                
                # copy daily variables template
                self.copy_daily_variables(os.path.join(self.source_path, file), date_str, data_count)

                self.logger(f"✅ Copied daily variables for file: {file}")

                # copy main template
                instance_template = self.copy_template(self.template_path, date_str, data_count)

                self.logger(f"✅ Copied template for file: {file}")

                # populate productivity and MPDM sheets in the copied template
                source_file_path = os.path.join(self.source_path, file)
                self._populate_productivity(source_file_path, instance_template)
                self._populate_mpdm(source_file_path, instance_template)

                # update progress
                progress = int(((index + 1) / len(total_files)) * 100)
                self.update_progress(progress)

        # For example, you can call other methods to read templates, process data, etc.
        self.logger("✅ Data processing completed!")

    def format_data_count(self, data_count=1):
        try:
            return f"{data_count:02d}"
        except Exception as e:
            self.logger(f"❌ Error formatting data count: {e}")
            return "00"

    def populate_productivity(self, source_file, template_path):
        try:
            source_wb = load_workbook(filename=source_file, read_only=False)

            # identify_equipment type
            if self.equipment.lower() in ['dozer']:
                self._populate_dozer_productivity(source_file, template_path)

            # source_ws = source_wb[self.equipment_sheet_name]
            source_ws = source_wb['dozer']

            new_wb = load_workbook(filename=template_path, read_only=False)
            # new_ws = new_wb[self.equipment_sheet_name.title()]
            new_ws = new_wb['Dozer']
          
            # Copy values and styles from source to new sheet
            project_code = source_ws['B7'].value
            number_of_equipment_types = source_ws['D7'].value
            date = source_ws['A7'].value
            data_collector = source_ws['C7'].value

            COLUMN_MAPPINGS = {
                'E': 'B',  # Equipment Tag (Dozer Cyle)
                'F': 'C',  # Man power
                'G': 'D',  # Dozer Blade Type
                'H': 'E',  # Task Type
                'I': 'F',  # Description
                'J': 'G',  # Soil Type
                'K': 'H',  # Blade Height (m)
                'L': 'I',  # Blade Width (m)
                'M': 'J',  # Blade Length (m)
                'N': 'K',  # unit (m3, m, etc)
                #'O': 'L',  # Blade Load (m3, m, etc) - This will be calculated, so we can skip copying this column
                'P': 'M',  # Cycle Time (seconds)
                # 'Q': 'N',  # Productivity (m3/hr, m/hr, etc) - This will be calculated, so we can skip copying this column
            }
            source_start_row = 7  # Assuming the first row is headers
            dest_start_row = 11  # Assuming the first row is headers
            

            # Copy project code, date, data collector, and number of equipment types to the new sheet
            new_ws['D6'] = project_code
            new_ws['I6'] = number_of_equipment_types
            new_ws['L6'] = date.strftime("%d-%m-%Y") if hasattr(date, "strftime") else date
            new_ws['O6'] = "Data Collector"
            new_ws['P6'] = data_collector

            
            # calculate how many rows down the data offset needs to shift
            row_offset = dest_start_row - source_start_row

            for row in range(source_start_row, source_ws.max_row + 1):

                source_value = source_ws[f"P{row}"].value

                # Optional skip empty rows
                if source_value is None:
                    continue

                target_row = row + row_offset

                self.logger(
                    f"Processing MPDM row {row}: {source_value}"
                )

                for source_col, target_col in COLUMN_MAPPINGS.items():

                    new_ws[f"{target_col}{target_row}"] = (
                        source_ws[f"{source_col}{row}"].value
                    )

            # new_wb.save(template_path)

            self.logger(
                f"✅ Updated Productivity sheet in: {template_path}"
            ) 
           
            dest_path = os.path.join(self.output_folder, os.path.basename(template_path))
            new_wb.save(dest_path)
            self.logger(f"✅ Updated productivity sheet in: {dest_path}")
            source_wb.close()
        except Exception as e:
            self.logger(f"❌ Error populating productivity sheet: {e}")

    def populate_mpdm(self, source_file_path, template_path):
        try:
            wb = load_workbook(filename=source_file_path, read_only=False)
            ws = wb['mpdm']
            ws['B2'] = "Sample MPDM Data"

            # dest_path = os.path.join(self.output_folder, os.path.basename(template_path))
            template_wb = load_workbook(filename=template_path, read_only=False)
            template_ws = template_wb['MPDM 1']
            template_start_row = 13
            source_start_row = 7

            for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=8, max_col=ws.max_column, values_only=False):
                # Grab the matching target cell in the new sheet
                source_cell = row[0].row 
                row_diff = template_start_row - source_start_row
                
                self.logger(f"Processing MPDM row {source_cell} with values: {ws['H'+str(source_cell)].value}")
                template_ws[f'B{row[0].row + row_diff}'] = ws[f"H{row[0].row}"].value 
                template_ws[f'V{row[0].row + row_diff}'] = ws[f"I{row[0].row}"].value 
                template_ws[f'W{row[0].row + row_diff}'] = ws[f"J{row[0].row}"].value 
                template_ws[f'X{row[0].row + row_diff}'] = ws[f"K{row[0].row}"].value 
                template_ws[f'Y{row[0].row + row_diff}'] = ws[f"L{row[0].row}"].value 
                template_ws[f'Z{row[0].row + row_diff}'] = ws[f"M{row[0].row}"].value 
                template_ws[f'AA{row[0].row + row_diff}'] = ws[f"N{row[0].row}"].value 

                # for cell in row:
                #     dest_cell = template_ws.cell(row=cell.row, column=cell.column)
                #     dest_cell.value = cell.value
                    
            # template_ws['N2'] = "Sample MPDM Data"
            template_wb.save(template_path)
            self.logger(f"✅ Updated MPDM sheet in: {template_path}")
            wb.close()
        except Exception as e:
            self.logger(f"❌ Error populating MPDM sheet: {e}")

    def _populate_productivity(self, source_file_path, template_path):
        try:
            # Identify equipment here
            equipment, config = self.get_equipment_config()

            source_wb = load_workbook(source_file_path)
            template_wb = load_workbook(template_path)

            source_ws = source_wb[
                config["source_sheet"]
            ]

            template_ws = template_wb[
                config["destination_sheet"]
            ]

            source_start_row = config["source_start_row"]

            dest_start_row = config["dest_start_row"]

            row_offset = (dest_start_row - source_start_row)

            column_mappings = config["column_mappings"]

            self._copy_header_data(source_ws, template_ws, config)


            for row in range(source_start_row, source_ws.max_row + 1):
                target_row = row + row_offset
                for source_col, dest_col in column_mappings.items():
                    template_ws[f"{dest_col}{target_row}"] = source_ws[f"{source_col}{row}"].value

                template_wb.save(template_path)

            # if "meta_data" in config["header_mappings"]:
            #     pass
            self.logger(
                f"✅ Successfully populated {equipment}"
            )

            source_wb.close()
            template_wb.close()
       

        except Exception as e:
            self.logger(f"❌ Error populating productivity: {e}")

    def _copy_header_data(self, source_ws, template_ws, config):
        header_config = config.get("header_mappings", {})
        if not header_config:
            return 

        date = header_config.get("date")
        project_code = header_config.get("project_code")
        data_collector = header_config.get("data_collector")
        number_of_equipment_types = header_config.get("number_of_equipment_types")
        if date:
            template_ws['L6'] = source_ws[date].value
        if project_code:
            template_ws['M6'] = source_ws[project_code].value
        if data_collector:
            template_ws['N6'] = source_ws[data_collector].value
        if number_of_equipment_types:
            template_ws['O6'] = source_ws[number_of_equipment_types].value


    def _populate_mpdm(self, source_file_path, template_path):
        try:
            source_wb = load_workbook(source_file_path)
            source_ws = source_wb["mpdm"]

            template_wb = load_workbook(template_path)
            template_ws = template_wb["MPDM 1"]

            source_start_row = 7
            template_start_row = 13
            row_offset = template_start_row - source_start_row

            # Source -> Destination mapping
            column_mapping = {
                "H": "B",
                "I": "V",
                "J": "W",
                "K": "X",
                "L": "Y",
                "M": "Z",
                "N": "AA",
            }

            for row in range(source_start_row, source_ws.max_row + 1):

                source_value = source_ws[f"H{row}"].value

                # Optional skip empty rows
                if source_value is None:
                    continue

                target_row = row + row_offset

                # self.logger(
                #     f"Processing MPDM row {row}: {source_value}"
                # )

                for source_col, target_col in column_mapping.items():

                    template_ws[f"{target_col}{target_row}"] = (
                        source_ws[f"{source_col}{row}"].value
                    )

            template_wb.save(template_path)

            self.logger(
                f"✅ Updated MPDM sheet in: {template_path}"
            )

            source_wb.close()
            template_wb.close()

        except Exception as e:
            self.logger(f"❌ Error populating MPDM sheet: {e}")

    def get_sheet_by_flexible_name(self, wb, target_names):
        """
        Find sheet using flexible matching.
        """

        # normalize targets
        normalized_targets = [
            name.lower()
                .replace("_", "")
                .replace(" ", "")
            for name in target_names
        ]

        for sheet_name in wb.sheetnames:

            normalized_sheet = (
                sheet_name.lower()
                    .replace("_", "")
                    .replace(" ", "")
            )

            if normalized_sheet in normalized_targets:
                return wb[sheet_name]

        return None
    def copy_daily_variables(self, file_path, date_str, data_count):

        try:
            # daily_variables_template = os.path.join(self.template_path.parent, "daily_variables_BiT.xlsx")
            wb = load_workbook(filename=file_path, read_only=False)
            ws = ws = self.get_sheet_by_flexible_name(
                wb, 
                self.SHEET_ALIASES["daily_variables"]
            )
                
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = "daily_variables"
            for row in ws.iter_rows(values_only=False):
                # Grab the matching target cell in the new sheet
                for cell in row:
                    new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                
                # deep copy structural attributes
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

            # 2. Copy Column Dimensions (Widths)
            for col_letter, col_dim in ws.column_dimensions.items():
                new_ws.column_dimensions[col_letter].width = col_dim.width
                new_ws.column_dimensions[col_letter].hidden = col_dim.hidden

            # 3. Copy Row Dimensions (Heights)
            for row_idx, row_dim in ws.row_dimensions.items():
                new_ws.row_dimensions[row_idx].height = row_dim.height
                new_ws.row_dimensions[row_idx].hidden = row_dim.hidden

            # 4. Copy Merged Cell Ranges (if any exist)
            for merged_range in list(ws.merged_cells.ranges):
                new_ws.merge_cells(str(merged_range))    
                # new_ws.append([cell.value for cell in row])

            # Finally update data count and date in the new sheet
            new_ws['E5'] = data_count

            
            # Save logic pipeline
            dest_folder = os.path.join(self.source_path, "daily_variables")
            if not os.path.exists(dest_folder):
                os.makedirs(dest_folder)
                
            dest_path = os.path.join(dest_folder, f"{data_count}_{date_str}_daily_variables.xlsx")
            new_wb.save(dest_path)
            wb.close()  # Clean up file stream locks
            
            self.logger(f"✅ Copied daily variables with complete styles to: {dest_path}")
        except Exception as e:
            self.logger(f"❌ Parsing Error: {str(e)}")


    def copy_template(self, template_path, date_str=None, data_count="00"):  

        dest_folder = os.path.join(self.source_path, "output")
        try:
            if not os.path.exists(dest_folder):
                os.makedirs(dest_folder)
                self.logger(f"📁 Created folder: {dest_folder}")
            
            # dest_path = os.path.join(dest_folder, os.path.basename(self.source_path))
            # dest_path = os.path.join(dest_folder, f"{data_count}_{date_str}_{os.path.basename(template_path)}")
            dest_path = os.path.join(dest_folder, f"{data_count}_{date_str}_template.xlsx")

            shutil.copy(template_path, dest_path)
            self.logger(f"✅ Copied template to: {dest_path}")
            return dest_path
          
        except Exception as e:
            self.logger(f"❌ Error copying template: {e}")
            return None

    def get_date_str(self, file_path):
        self.logger(f"reading date: from {file_path}")
        try:
            
            file_name = os.path.basename(file_path)
            mach = re.search(self.date_pattern, file_name)
            if mach:
                return mach.group(1)


            wb = load_workbook(file_path, data_only=True)
            ws = wb['daily_variables']
            date_value = ws['H5'].value
            if hasattr(date_value, "strftime"):
                return date_value.strftime("%d-%m-%Y")
            
            else:
                if "/" in str(date_value):
                    return date_value.replace("/", "-")
                if "_" in str(date_value):
                    return date_value.replace("_", "-")
                return date_value

        except Exception as e:
            self.logger(f"❌ Error reading date from {file_path}: {e}")
            return None
        
    def read_template_file(self, template_path):
        try:
            wb = load_workbook(filename=template_path, read_only=True)
            self.logger(f"✅ Successfully loaded template: {template_path}")
            return wb
        except Exception as e:
            self.logger(f"❌ Error loading template: {e}")
            return None

    def rename_template_file(self, template_path, new_name):
        try:
            dest_path = os.path.join(self.output_folder, new_name)
            shutil.copy(template_path, dest_path)
            self.logger(f"✅ Renamed template to: {dest_path}")
            return dest_path
        except Exception as e:
            self.logger(f"❌ Error renaming template: {e}")
            return None
        
    
        
    def load_source_data(self, file_path):
        wb = load_workbook(filename=file_path, read_only=True)

        return wb 