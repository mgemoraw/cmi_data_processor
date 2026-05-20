from openpyxl import load_workbook
import shutil 
import os
import sys
from pathlib import  Path 
import re 



class FileRenamingEngine:
    date_pattern =  r"(\d{2}-\d{2}-\d{4})"


    def __init__(self, input_folder=None, file_type=".xlsx", logger=None):
        self.file_type= file_type
        self.input_folder = Path(input_folder)
        self.logger = logger
        self.progress_callback = None
        
        self.output_folder = os.path.join(self.input_folder, "renamed_files")

        # Pattern for DD-MM-YYYY
        self.date_pattern = r"(\d{2}-\d{2}-\d{4})"

    def _log(self, message):
        """Sends message to GUI log or console"""
        if self.logger:
            self.logger(message)
        else:
            print(message)

    def read_excel_contents(self):
        """Main execution method called by the Worker thread"""
        
        # 1. Setup Destination
        dest_folder = os.path.join(self.input_folder, "renamed_files")
        if not os.path.exists(dest_folder):
            os.makedirs(dest_folder)
            self._log(f"📁 Created folder: {dest_folder}")

        # 2. Filter Files
        files = [f for f in os.listdir(self.input_folder) 
                 if f.lower().endswith(self.file_type.lower())]
        total_files = len(files)

        if total_files == 0:
            self._log(f"⚠️ No {self.file_type} files found in source.")
            return

        # 3. Process loop
        try:
            for index, filename in enumerate(files):
                src_path = os.path.join(self.input_folder, filename)
                
                # Extract Date
                match = re.search(self.date_pattern, filename)
                
                if match:
                    date_str = match.group(1)
                    
                    # Logic: Move date to the front, clean up double underscores
                    clean_name = filename.replace(date_str, "").replace("__", "_").strip("_")
                    new_filename = f"{date_str}_{clean_name}"
                    
                    dest_path = os.path.join(dest_folder, new_filename)
                    
                    # Copy the file
                    shutil.copy2(src_path, dest_path)
                    self._log(f"✅ Renamed: {filename} -> {new_filename}")
                else:
                    date_str = self.get_date_str(src_path)
                    if not date_str:
                        date_str = self.get_date_str_from_mpdm(src_path)
                    
                    if date_str:
                        new_filename = f"{date_str}_{filename}"
                        dest_path = os.path.join(dest_folder, new_filename)

                        self._log(f"Destination: {dest_path}")
                        shutil.copy2(src_path, dest_path)
                        self._log(f"✅ Renamed: {filename} -> {new_filename}")
                    else:
                        self._log(f"ℹ️ Skipping: {filename} (No date found)")

                # 4. Update Progress Bar in GUI
                if self.progress_callback:
                    progress_value = int(((index + 1) / total_files) * 100)
                    self.progress_callback(progress_value)
        except Exception as e:
            self._log(f"❌ Error processing files: {e}")


    def start(self):
        # if destination folder doesn't exist create it
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)
            print(f"Created folder: {self.output_folder}")


        # self.read_files(self.input_folder)
        self.process_naming()

    def rename(self, folder_path, filename, date_str):
        
        new_filename = f"{date_str}_" + filename
        # new_file = os.path.join(folder_path, new_filename)
        return new_filename

    def get_date_str(self, file_path):
        print(f"reading date: from {file_path}")
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb['daily-variables']


            date_value = ws['G4'].value
            if hasattr(date_value, "strftime"):
                return date_value.strftime("%d-%m-%Y")
            
            else:
                if "/" in str(date_value):
                    return date_value.replace("/", "-")
                return date_value

        except Exception as e:
            print(f"Error rading {file_path}: {e}")
            return None
        
    def get_date_str_from_mpdm(self, file_path):
        print(f"reading date: from {file_path}")
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb['mpdm']


            date_value = ws['J6'].value
            if hasattr(date_value, "strftime"):
                return date_value.strftime("%d-%m-%Y")
            
            else:
                if "/" in str(date_value):
                    return date_value.replace("/", "-")
                return date_value

        except Exception as e:
            print(f"Error rading {file_path}: {e}")
            return None
    
    def process_naming(self):
        for filename in os.listdir(self.input_folder):
            source_path = os.path.join(self.input_folder, filename)
            print(source_path)

            # Only proces files 
            if os.path.isfile(source_path):
                # rename 
                match = re.search(self.date_pattern, filename)
                if match:
                    # copy the file to the new location
                    destination_path = os.path.join(self.output_folder, filename)

                    shutil.copy2(source_path, destination_path)
                    print(f"Saved: {filename} -> {destination_path}/{filename}")
                    continue
                
                date_str = self.get_date_str(source_path)
                if not date_str:
                    raise ValueError("None is not a valid date string")

                new_name = self.rename(self.input_folder, filename, date_str)

                destination_path = os.path.join(self.output_folder, new_name)
                print("Destination: ", destination_path)

                # copy the file to the new location
                shutil.copy2(source_path, destination_path)
                print(f"Saved: {filename} -> {destination_path}/{new_name}")

    def read_files(self, folder_path):
        folder_path = Path(folder_path)


        # Iterate through files
        for filename in os.listdir(folder_path):
            # Construct the full file paths
            old_file = os.path.join(folder_path, filename)


            # Skip if it's a folder (we only want files)
            if os.path.isfile(old_file):
                date_str = self.get_date_str(filename)
                new_file = self.rename(folder_path, filename, date_str)


                # rename and save
                os.rename(old_file, new_file)
                print(f"Renamed: {filename} -> {new_file}")


if __name__ == "__main__":
    rn = FileRenamingEngine(
        input_folder="./example/",
        file_type=".xlsx",
    )

    rn.start()