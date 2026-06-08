import os
from pathlib import Path
from openpyxl import load_workbook

class RowFillEngine:
    def __init__(self, input_folder, logger=None, progress_callback=None):
        self.source_folder = Path(input_folder)
        self.logger = logger
        self.progress_callback = progress_callback

    def _log(self, text):
        if self.logger:
            self.logger(text)
        else:
            print(text)

    def _set_progress(self, val):
        if self.progress_callback:
            self.progress_callback(val)

    def read_excel_contents(self):
        """Standard execution hook matching your GUI Worker class thread setup."""
        self.run()

    def run(self):
        self._log("🏁 Starting Row Fill Blank Optimization Engine...")
        
        if not self.source_folder.exists() or not self.source_folder.is_dir():
            self._log("❌ Error: Target input directory does not exist.")
            return False

        all_files = [f for f in os.listdir(self.source_folder) if f.lower().endswith(".xlsx")]
        total_files = len(all_files)

        if total_files == 0:
            self._log("⚠️ No active spreadsheet data logs located inside target input folder directory.")
            return False

        for index, file in enumerate(all_files):
            full_path = self.source_folder / file
            self._log(f"🔄 Analyzing sheet boundaries for data persistence inside: {file}")

            try:
                # Open workbook with data_only=False to preserve any existing formulas
                wb = load_workbook(filename=str(full_path), data_only=False)
                ws = wb.active  # Processes the active layout tab segment
                
                # Context Range Boundaries matching Data Analysis (Rows 11 to 111)
                start_row = 11
                end_row = 111
                
                # 1. First, find the very first valid row containing data to use as a baseline template
                baseline_values = {}
                first_data_row_idx = None

                for row in range(start_row, end_row + 1):
                    row_has_data = any(ws.cell(row=row, column=col).value is not None for col in range(1, ws.max_column + 1))
                    if row_has_data:
                        first_data_row_idx = row
                        # Cache the baseline cell entries
                        for col in range(1, ws.max_column + 1):
                            baseline_values[col] = ws.cell(row=row, column=col).value
                        break

                if first_data_row_idx is None:
                    self._log(f"⚠️ No structural records found inside sheet layout rows {start_row}-{end_row} for {file}. Skipping.")
                    wb.close()
                    continue

                self._log(f"📋 Baseline reference structure recognized at Row {first_data_row_idx}. Propagating empty cells...")

                # 2. Update subsequent blank cells using the cached baseline matrix
                cells_updated_count = 0
                for row in range(first_data_row_idx + 1, end_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        # If cell is empty, fill it with the baseline row's corresponding column item
                        if cell.value is None or str(cell.value).strip() == "":
                            cell.value = baseline_values.get(col, None)
                            cells_updated_count += 1

                wb.save(str(full_path))
                wb.close()
                self._log(f"✅ Filled ({cells_updated_count}) structural empty data properties successfully in: {file}")

            except Exception as ex:
                self._log(f"❌ Structural manipulation exception on file sequence: {file}. Details: {ex}")

            # Safe UI progress calculation update
            self._set_progress(int(((index + 1) / total_files) * 100))

        self._log("💾 All target documents filled and saved successfully.")
        return True