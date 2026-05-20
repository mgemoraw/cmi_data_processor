from openpyxl import load_workbook, Workbook
from pathlib import Path
from openpyxl.styles import (
	Font,
	Border,
	Side,
	Alignment,
	PatternFill
)
from copy import copy



class EquipmentProductivity:
	def __init__(self, folder_path="./", output_folder="output", equipment=None, logger=None):
		self.folder_path = Path(folder_path)
		self.output_folder = Path(output_folder)
		self.equipment = equipment
		self.logger = logger or print

		self.progress_callback = None

		# Create output folder if it does not exist
		self.output_folder.mkdir(exist_ok=True)

		# Header row and data start row
		self.header_row = 6
		self.data_start_row = 7

		# Split size
		self.chunk_size = 100

	def log(self, message):
		self.logger(message)


	def update_progress(self, value):
		if self.progress_callback:
			self.progress_callback(value)


	def list_excel_files(self):
		"""
		Return all Excel files in the folder
		"""
		return list(self.folder_path.glob("*.xlsx"))

	def read_excel_contents(self):

		files = self.list_excel_files()
		# to update progress
		total_files = len(files)


		if not files:
			# print("No Excel files found.")
			self.log("No Excel files found.")
			return

		
		for index, file in enumerate(files):

			# print(f"\nReading file: {file.name}")
			self.log(f"\nReading file: {file.name}")

			try:
				wb = load_workbook(file, data_only=True)

				# if "excavator" in wb.sheetnames:
				if self.equipment in wb.sheetnames:
					# self.excavator_splitter(wb, file.stem)
					self.productivity_splitter(wb, file.stem)
			
				progress = int(((index + 1) / total_files) * 100)
				self.update_progress(progress)

			except Exception as e:
				# print(f"Error reading {file.name}: {e}")
				self.log(f"Error reading {file.name}: {e}")

	def copy_template(
		self,
		source_ws,
		target_ws
	):
		"""
		Copy rows before/including header,
		merged cells, row heights, and
		column widths.
		"""

		# --------------------------------
		# Copy Rows (1 -> header row)
		# --------------------------------
		for row in source_ws.iter_rows(
			min_row=1,
			max_row=self.header_row
		):

			for cell in row:

				new_cell = target_ws[cell.coordinate]

				# Copy value
				new_cell.value = cell.value

				# Copy styles
				if cell.has_style:

					new_cell.font = copy(cell.font)
					new_cell.fill = copy(cell.fill)
					new_cell.border = copy(cell.border)
					new_cell.alignment = copy(cell.alignment)
					new_cell.number_format = copy(
						cell.number_format
					)
					new_cell.protection = copy(
						cell.protection
					)

		# --------------------------------
		# Copy Merged Cells
		# --------------------------------
		for merged_range in source_ws.merged_cells.ranges:

			if merged_range.min_row <= self.header_row:

				target_ws.merge_cells(
					str(merged_range)
				)

		# --------------------------------
		# Copy Row Heights
		# --------------------------------
		for row_num in range(
			1,
			self.header_row + 1
		):

			source_dimension = (
				source_ws.row_dimensions[row_num]
			)

			target_ws.row_dimensions[
				row_num
			].height = source_dimension.height

		# --------------------------------
		# Copy Column Widths
		# --------------------------------
		for col_letter, dimension in (
			source_ws.column_dimensions.items()
		):

			target_ws.column_dimensions[
				col_letter
			].width = dimension.width

	def copy_sheet_styles(
		self,
		source_ws,
		target_ws
	):
		"""
		Copy worksheet-level settings.
		"""

		# --------------------------------
		# Freeze Panes
		# --------------------------------
		target_ws.freeze_panes = (
			source_ws.freeze_panes
		)

		# --------------------------------
		# Sheet View
		# --------------------------------
		target_ws.sheet_view.zoomScale = (
			source_ws.sheet_view.zoomScale
		)

		# --------------------------------
		# Auto Filter
		# --------------------------------
		if source_ws.auto_filter:

			target_ws.auto_filter.ref = (
				source_ws.auto_filter.ref
			)

		# --------------------------------
		# Sheet Properties
		# --------------------------------
		target_ws.sheet_format.defaultRowHeight = (
			source_ws.sheet_format.defaultRowHeight
		)

		# --------------------------------
		# Page Margins
		# --------------------------------
		target_ws.page_margins = copy(
			source_ws.page_margins
		)

		# --------------------------------
		# Page Setup
		# --------------------------------
		target_ws.page_setup = copy(
			source_ws.page_setup
		)

		# --------------------------------
		# Print Options
		# --------------------------------
		target_ws.print_options = copy(
			source_ws.print_options
		)

		# --------------------------------
		# Sheet Properties
		# --------------------------------
		target_ws.sheet_properties = copy(
			source_ws.sheet_properties
		)

	def write_data_rows(
		self,
		source_ws,
		target_ws,
		rows_data
	):
		"""
		Write data rows starting at
		original data start row while
		preserving styles.
		"""

		start_row = self.data_start_row

		for row_data in rows_data:

			for col_index, value in enumerate(
				row_data,
				start=1
			):

				# Create target cell
				target_cell = target_ws.cell(
					row=start_row,
					column=col_index,
					value=value
				)

				# Reference style from
				# original first data row
				source_cell = source_ws.cell(
					row=self.data_start_row,
					column=col_index
				)

				# Copy styles
				if source_cell.has_style:

					target_cell.font = copy(
						source_cell.font
					)

					target_cell.fill = copy(
						source_cell.fill
					)

					target_cell.border = copy(
						source_cell.border
					)

					target_cell.alignment = copy(
						source_cell.alignment
					)

					target_cell.number_format = copy(
						source_cell.number_format
					)

					target_cell.protection = copy(
						source_cell.protection
					)

			start_row += 1

	def mpdm_splitter(
		self,
		source_ws,
		target_ws,
		allowed_cycles,
	):
		"""
		Split MPDM sheet based on allowed cycle numbers
		from excavator chunk.
		"""

		# --------------------------------
		# Read headers (Row 6)
		# --------------------------------
		headers = [
			cell.value
			for cell in source_ws[self.header_row]
		]

		# Find cycle column index
		cycle_col = headers.index(
			"Production Cycle Number"
		)

		# --------------------------------
		# Copy Template (Rows 1-6)
		# --------------------------------
		for row in source_ws.iter_rows(
			min_row=1,
			max_row=self.header_row
		):

			for cell in row:

				new_cell = target_ws[cell.coordinate]

				new_cell.value = cell.value

				if cell.has_style:
					new_cell.font = copy(cell.font)
					new_cell.fill = copy(cell.fill)
					new_cell.border = copy(cell.border)
					new_cell.alignment = copy(cell.alignment)
					new_cell.number_format = copy(
						cell.number_format
					)
					new_cell.protection = copy(cell.protection)

		# --------------------------------
		# Copy merged cells (only header area)
		# --------------------------------
		for merged_range in source_ws.merged_cells.ranges:

			if merged_range.min_row <= self.header_row:

				target_ws.merge_cells(str(merged_range))

		# --------------------------------
		# Copy column widths
		# --------------------------------
		for col_letter, dim in source_ws.column_dimensions.items():

			target_ws.column_dimensions[
				col_letter
			].width = dim.width

		# --------------------------------
		# Write filtered MPDM data
		# --------------------------------
		start_row = self.data_start_row

		for row in source_ws.iter_rows(
			min_row=self.data_start_row,
			values_only=True
		):

			if all(v is None for v in row):
				continue

			cycle_value = row[cycle_col]

			if cycle_value not in allowed_cycles:
				continue

			for col_index, value in enumerate(row, start=1):

				cell = target_ws.cell(
					row=start_row,
					column=col_index,
					value=value
				)

				# Copy style from original data row
				source_cell = source_ws.cell(
					row=self.data_start_row,
					column=col_index
				)

				if source_cell.has_style:

					cell.font = copy(source_cell.font)
					cell.fill = copy(source_cell.fill)
					cell.border = copy(source_cell.border)
					cell.alignment = copy(source_cell.alignment)
					cell.number_format = copy(
						source_cell.number_format
					)
					cell.protection = copy(
						source_cell.protection
					)

			start_row += 1

		# --------------------------------
		# Freeze header
		# --------------------------------
		target_ws.freeze_panes = f"A{self.data_start_row}"

		# --------------------------------
		# Auto filter
		# --------------------------------
		# target_ws.auto_filter.ref = target_ws.dimensions


	def get_chunk_cycles(self, chunk, cycle_col_index):
		"""
		Extract unique cycle numbers from a chunk
		of excavator data.
		"""

		cycles = set()

		for row in chunk:

			if not row:
				continue

			cycle_value = row[cycle_col_index]

			if cycle_value is not None:
				cycles.add(cycle_value)

		return cycles


	def productivity_splitter(self, workbook, original_filename):
		if self.equipment is None:
			raise ValueError("Equipment cannot be Empty")

		
		ws = workbook[self.equipment]
		mpdm_ws = workbook['mpdm']


		# ----------------------------
		# Read headers from row 6
		# ----------------------------
		headers = []

		for cell in ws[self.header_row]:
			headers.append(cell.value)

		# ----------------------------
		# Read data starting from row 7
		# ----------------------------
		data_rows = []

		for row in ws.iter_rows(
			min_row=self.data_start_row,
			values_only=True
		):

			if all(value is None for value in row):
				continue

			data_rows.append(row)

		# print(f"\nTotal Data Rows: {len(data_rows)}")
		self.log(f"\nTotal Data Rows: {len(data_rows)}")

		# ----------------------------
		# Styles
		# ----------------------------

		thin = Side(
			border_style="thin",
			color="000000"
		)

		border = Border(
			left=thin,
			right=thin,
			top=thin,
			bottom=thin
		)

		header_font = Font(
			bold=True,
			color="FFFFFF"
		)

		header_fill = PatternFill(
			fill_type="solid",
			start_color="1F4E78"
		)

		header_alignment = Alignment(
			horizontal="center",
			vertical="center",
			wrap_text=True
		)

		body_alignment = Alignment(
			vertical="top",
			wrap_text=True
		)

		# ----------------------------
		# Split into chunks
		# ----------------------------

		for chunk_index in range(0, len(data_rows), self.chunk_size):

			chunk = data_rows[
				chunk_index:chunk_index + self.chunk_size
			]

			# Create new workbook
			new_wb = Workbook()
			new_ws = new_wb.active
			new_ws.title = self.equipment

			# =====================
			# CREATE MPDM SHEET
			# =====================
			mpdm_new_ws = new_wb.create_sheet("mpdm")
			exc_headers = [
				cell.value
				for cell in ws[self.header_row]
			]

			cycle_col = f"{self.equipment.title()} Cycle"
			cycle_col = headers.index(cycle_col)

			# ---------------------------
			# COLLECT CHUNK CYCLES AND SPLIT MPDM
			# ---------------------------
			chunk_cycles = self.get_chunk_cycles(chunk, cycle_col)
			self.mpdm_splitter(
				mpdm_ws,
				mpdm_new_ws,
				chunk_cycles
			)

			# ==========================
			# copy daily variables to each chunk
			# ==========================
			new_dv_ws = workbook.create_sheet('daily_variables')
			# self.copy_daily_variables(workbook, new_wb)
			self.clone_sheet_exact(workbook, new_wb)


			# ----------------------------
			# Copy top section (Rows 1-6)
			# ----------------------------

			for row in ws.iter_rows(
				min_row=1,
				max_row=self.header_row
			):

				for cell in row:

					new_cell = new_ws[cell.coordinate]

					# Copy value
					new_cell.value = cell.value

					# Copy style
					if cell.has_style:
						new_cell.font = copy(cell.font)
						new_cell.fill = copy(cell.fill)
						new_cell.border = copy(cell.border)
						new_cell.alignment = copy(cell.alignment)
						new_cell.number_format = copy(cell.number_format)
						new_cell.protection = copy(cell.protection)

			# ----------------------------
			# Copy merged cells
			# ----------------------------

			for merged_range in ws.merged_cells.ranges:

				if merged_range.min_row <= self.header_row:
					new_ws.merge_cells(str(merged_range))

			# ----------------------------
			# Copy row heights
			# ----------------------------

			for row_num in range(1, self.header_row + 1):

				if row_num in ws.row_dimensions:

					new_ws.row_dimensions[row_num].height = (
						ws.row_dimensions[row_num].height
					)

			# ----------------------------
			# Copy column widths
			# ----------------------------

			for col_letter, dimension in ws.column_dimensions.items():

				new_ws.column_dimensions[col_letter].width = (
					dimension.width
				)

			# ----------------------------
			# Write chunk data starting at row 7
			# ----------------------------

			start_row = self.data_start_row

			for row_data in chunk:

				for col_index, value in enumerate(row_data, start=1):

					cell = new_ws.cell(
						row=start_row,
						column=col_index,
						value=value
					)

					# Copy style from original row 7
					source_cell = ws.cell(
						row=self.data_start_row,
						column=col_index
					)

					if source_cell.has_style:

						cell.font = copy(source_cell.font)
						cell.fill = copy(source_cell.fill)
						cell.border = copy(source_cell.border)
						cell.alignment = copy(source_cell.alignment)
						cell.number_format = copy(source_cell.number_format)
						cell.protection = copy(source_cell.protection)

				start_row += 1



			# Freeze header
			new_ws.freeze_panes = f"A{self.data_start_row}"

			# Auto filter
			# new_ws.auto_filter.ref = new_ws.dimensions

			# ----------------------------
			# Save file
			# ----------------------------

			file_number = (
				chunk_index // self.chunk_size
			) + 1

			output_file = (
				self.output_folder /
				f"{original_filename}_part_{file_number}.xlsx"
			)

			new_wb.save(output_file)

			# print(f"Saved: {output_file}")
			self.log(f"Saved: {output_file}")


	def excavator_splitter(self, workbook, original_filename):

		ws = workbook["excavator"]
		mpdm_ws = workbook['mpdm']


		# ----------------------------
		# Read headers from row 6
		# ----------------------------
		headers = []

		for cell in ws[self.header_row]:
			headers.append(cell.value)

		# ----------------------------
		# Read data starting from row 7
		# ----------------------------
		data_rows = []

		for row in ws.iter_rows(
			min_row=self.data_start_row,
			values_only=True
		):

			if all(value is None for value in row):
				continue

			data_rows.append(row)

		# print(f"\nTotal Data Rows: {len(data_rows)}")
		self.log(f"\nTotal Data Rows: {len(data_rows)}")

		# ----------------------------
		# Styles
		# ----------------------------

		thin = Side(
			border_style="thin",
			color="000000"
		)

		border = Border(
			left=thin,
			right=thin,
			top=thin,
			bottom=thin
		)

		header_font = Font(
			bold=True,
			color="FFFFFF"
		)

		header_fill = PatternFill(
			fill_type="solid",
			start_color="1F4E78"
		)

		header_alignment = Alignment(
			horizontal="center",
			vertical="center",
			wrap_text=True
		)

		body_alignment = Alignment(
			vertical="top",
			wrap_text=True
		)

		# ----------------------------
		# Split into chunks
		# ----------------------------

		for chunk_index in range(0, len(data_rows), self.chunk_size):

			chunk = data_rows[
				chunk_index:chunk_index + self.chunk_size
			]

			# Create new workbook
			new_wb = Workbook()
			new_ws = new_wb.active
			new_ws.title = "excavator"

			# =====================
			# CREATE MPDM SHEET
			# =====================
			mpdm_new_ws = new_wb.create_sheet("mpdm")
			exc_headers = [
				cell.value
				for cell in ws[self.header_row]
			]
			exc_cycle_col = headers.index("Excavator Cycle")

			# ---------------------------
			# COLLECT CHUNK CYCLES AND SPLIT MPDM
			# ---------------------------
			chunk_cycles = self.get_chunk_cycles(chunk, exc_cycle_col)
			self.mpdm_splitter(
				mpdm_ws,
				mpdm_new_ws,
				chunk_cycles
			)

			# ==========================
			# copy daily variables to each chunk
			# ==========================
			self.copy_daily_variables(workbook, new_wb)


			# ----------------------------
			# Copy top section (Rows 1-6)
			# ----------------------------

			for row in ws.iter_rows(
				min_row=1,
				max_row=self.header_row
			):

				for cell in row:

					new_cell = new_ws[cell.coordinate]

					# Copy value
					new_cell.value = cell.value

					# Copy style
					if cell.has_style:
						new_cell.font = copy(cell.font)
						new_cell.fill = copy(cell.fill)
						new_cell.border = copy(cell.border)
						new_cell.alignment = copy(cell.alignment)
						new_cell.number_format = copy(cell.number_format)
						new_cell.protection = copy(cell.protection)

			# ----------------------------
			# Copy merged cells
			# ----------------------------

			for merged_range in ws.merged_cells.ranges:

				if merged_range.min_row <= self.header_row:
					new_ws.merge_cells(str(merged_range))

			# ----------------------------
			# Copy row heights
			# ----------------------------

			for row_num in range(1, self.header_row + 1):

				if row_num in ws.row_dimensions:

					new_ws.row_dimensions[row_num].height = (
						ws.row_dimensions[row_num].height
					)

			# ----------------------------
			# Copy column widths
			# ----------------------------

			for col_letter, dimension in ws.column_dimensions.items():

				new_ws.column_dimensions[col_letter].width = (
					dimension.width
				)

			# ----------------------------
			# Write chunk data starting at row 7
			# ----------------------------

			start_row = self.data_start_row

			for row_data in chunk:

				for col_index, value in enumerate(row_data, start=1):

					cell = new_ws.cell(
						row=start_row,
						column=col_index,
						value=value
					)

					# Copy style from original row 7
					source_cell = ws.cell(
						row=self.data_start_row,
						column=col_index
					)

					if source_cell.has_style:

						cell.font = copy(source_cell.font)
						cell.fill = copy(source_cell.fill)
						cell.border = copy(source_cell.border)
						cell.alignment = copy(source_cell.alignment)
						cell.number_format = copy(source_cell.number_format)
						cell.protection = copy(source_cell.protection)

				start_row += 1



			# Freeze header
			new_ws.freeze_panes = f"A{self.data_start_row}"

			# Auto filter
			# new_ws.auto_filter.ref = new_ws.dimensions

			# ----------------------------
			# Save file
			# ----------------------------

			file_number = (
				chunk_index // self.chunk_size
			) + 1

			output_file = (
				self.output_folder /
				f"{original_filename}_part_{file_number}.xlsx"
			)

			new_wb.save(output_file)

			# print(f"Saved: {output_file}")
			self.log(f"Saved: {output_file}")

	# =====================
	# CREATE DAILY VARIABLES SHEET
	# =====================
	def copy_daily_variables(self, source_wb, new_wb):

		dv = source_wb['daily_variables']
		dv_ws = new_wb.create_sheet("daily_variables")


		self.copy_template(
			dv,          # source sheet from original workbook
			dv_ws
		)

		self.copy_sheet_styles(
			dv,
			dv_ws
		)

		self.write_data_rows(
			dv,
			dv_ws,
			list(dv.iter_rows(
				min_row=self.data_start_row,
				values_only=True
			))
		)

	def clone_sheet_exact(self, source_wb, target_wb):

		source_ws = source_wb["daily_variables"]
		target_ws = target_wb.create_sheet("daily_variables")

		# =====================================
		# 1. COPY CELLS (VALUE + STYLE)
		# =====================================
		for row in source_ws.iter_rows():

			for cell in row:

				new_cell = target_ws.cell(
					row=cell.row,
					column=cell.column,
					value=cell.value
				)

				if cell.has_style:

					new_cell.font = copy(cell.font)
					new_cell.fill = copy(cell.fill)
					new_cell.border = copy(cell.border)
					new_cell.alignment = copy(cell.alignment)
					new_cell.number_format = copy(cell.number_format)
					new_cell.protection = copy(cell.protection)

		# =====================================
		# 2. COPY MERGED CELLS
		# =====================================
		for merged_range in source_ws.merged_cells.ranges:

			target_ws.merge_cells(str(merged_range))

		# =====================================
		# 3. COPY COLUMN WIDTHS
		# =====================================
		for col_letter, dim in source_ws.column_dimensions.items():

			if dim.width is not None:

				target_ws.column_dimensions[
					col_letter
				].width = dim.width

		# =====================================
		# 4. COPY ROW HEIGHTS
		# =====================================
		for row_num, dim in source_ws.row_dimensions.items():

			if dim.height is not None:

				target_ws.row_dimensions[
					row_num
				].height = dim.height

		# =====================================
		# 5. COPY SAFE SHEET SETTINGS ONLY
		# =====================================

		# Freeze panes (safe)
		target_ws.freeze_panes = source_ws.freeze_panes

		# Page margins (safe)
		target_ws.page_margins = copy(source_ws.page_margins)

		# Page setup (safe)
		target_ws.page_setup = copy(source_ws.page_setup)

		# Print options (safe)
		target_ws.print_options = copy(source_ws.print_options)

		# Sheet format (safe partial copy)
		target_ws.sheet_format.defaultRowHeight = (
			source_ws.sheet_format.defaultRowHeight
		)

		# =====================================
		# 6. AUTO FILTER (SAFE CHECK)
		# =====================================
		if (
			source_ws.auto_filter
			and source_ws.auto_filter.ref
		):
			target_ws.auto_filter.ref = source_ws.auto_filter.ref



if __name__ == "__main__":

	ep = ExcavatorProductivity(
		folder_path="./",
		output_folder=f"split_data_files",
		equipment="excavator",
	)

	ep.read_excel_contents()