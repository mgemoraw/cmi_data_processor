import sys
import os
import subprocess
from datetime import datetime
from PySide6.QtWidgets import (
    QFormLayout, QApplication, QWidget, QVBoxLayout,
    QPushButton, QFileDialog, QLineEdit, QLabel, 
    QTextEdit, QHBoxLayout, QComboBox, QProgressBar, QListWidget,
    QMenu  # Added for context menu construction
)
from PySide6.QtCore import QThread, Signal, Qt  # Added Qt namespace flags
from openpyxl import load_workbook

# =====================================================================
# THREAD WORKER - ASYNCHRONOUS METADATA EXTRACTOR
# =====================================================================
class ExcelMetaWorker(QThread):
    meta_loaded_signal = Signal(dict)
    error_signal = Signal(str)

    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path

    def run(self):
        try:
            stat_info = os.stat(self.file_path)
            file_size_kb = round(stat_info.st_size / 1024, 2)
            mod_time = datetime.fromtimestamp(stat_info.st_mtime).strftime('%Y-%m-%d %H:%M:%S')

            wb = load_workbook(self.file_path, read_only=True)
            sheet_names = wb.sheetnames
            num_sheets = len(sheet_names)
            
            ws = wb.active
            active_sheet_title = ws.title
            max_rows = ws.max_row if ws.max_row else "Unknown (Streaming Required)"
            max_cols = ws.max_column if ws.max_column else "Unknown"

            wb.close()

            metadata = {
                "file_size": f"{file_size_kb} KB",
                "mod_date": mod_time,
                "num_sheets": str(num_sheets),
                "sheet_names": ", ".join(sheet_names),
                "active_sheet": active_sheet_title,
                "row_count": str(max_rows),
                "col_count": str(max_cols)
            }
            self.meta_loaded_signal.emit(metadata)
        except Exception as e:
            self.error_signal.emit(str(e))


# =====================================================================
# THREAD WORKER - CORE ENGINES
# =====================================================================
class Worker(QThread):
    log_signal = Signal(str)
    progress_signal = Signal(int)
    finished_signal = Signal()

    def __init__(self, engine):
        super().__init__()
        self.engine = engine

    def run(self):
        self.engine.logger = self.log_signal.emit
        self.engine.progress_callback = self.progress_signal.emit
        if hasattr(self.engine, 'read_excel_contents'):
            self.engine.read_excel_contents()
        self.finished_signal.emit()


# =====================================================================
# MAIN WINDOW WITH METADATA SUMMARY VIEW
# =====================================================================
class CMIDataProcessorGUI(QWidget):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("CMI-Data Processor (Multi-Task Pro)")
        self.setMinimumSize(950, 700)
        self.init_ui()

    def init_ui(self):
        main_layout = QHBoxLayout()
        left_layout = QVBoxLayout()
        form = QFormLayout()

        # Task Selector
        self.task_selector = QComboBox()
        self.task_selector.addItems([
            "Excel File Splitting",
            "File Renaming (Date Match/G4 Extraction)",
            "Data Aggregation & Template Counter",
            "Clean Data"
        ])
        self.task_selector.currentIndexChanged.connect(self.toggle_task_inputs)
        form.addRow("Select Action Task:", self.task_selector)

        # Directory Fields
        input_widget = QWidget()
        input_layout = QHBoxLayout(input_widget)
        self.input_path = QLineEdit()
        self.input_path.textChanged.connect(self.populate_file_list)
        btn_input = QPushButton("Browse")
        btn_input.clicked.connect(self.select_input_folder)
        input_layout.addWidget(self.input_path)
        input_layout.addWidget(btn_input)
        form.addRow("Input Folder:", input_widget)

        output_widget = QWidget()
        output_layout = QHBoxLayout(output_widget)
        self.output_path = QLineEdit()
        btn_output = QPushButton("Browse")
        btn_output.clicked.connect(self.select_output_folder)
        output_layout.addWidget(self.output_path)
        output_layout.addWidget(btn_output)
        form.addRow("Output Folder:", output_widget)

        # Dynamic Parameters
        self.equipment = QComboBox()
        self.equipment.addItems(["Excavator", "Dozer", "Loader", "Grader", "Roller", "Truck", "Labor"])
        self.equipment_row_label = QLabel("Equipment Type:")
        form.addRow(self.equipment_row_label, self.equipment)

        self.chunk_size = QLineEdit()
        self.chunk_size.setPlaceholderText("100")
        self.chunk_row_label = QLabel("Chunk Size:")
        form.addRow(self.chunk_row_label, self.chunk_size)

        self.template_path_field = QLineEdit()
        self.template_path_field.setPlaceholderText("Path to baseline template .xlsx file...")
        self.template_btn = QPushButton("Browse Template")
        self.template_btn.clicked.connect(self.select_template_file)
        
        self.template_widget = QWidget()
        t_lay = QHBoxLayout(self.template_widget)
        t_lay.addWidget(self.template_path_field)
        t_lay.addWidget(self.template_btn)
        t_lay.setContentsMargins(0,0,0,0)
        
        self.template_row_label = QLabel("Template File:")
        form.addRow(self.template_row_label, self.template_widget)

        left_layout.addLayout(form)

        # Action Buttons
        button_layout = QHBoxLayout()
        self.start_btn = QPushButton("🚀 Start Process")
        self.start_btn.clicked.connect(self.execute_current_task)
        self.start_btn.setStyleSheet("background-color: #10b981; min-height: 38px;")
        self.open_output_btn = QPushButton("📂 Open Output Location")
        self.open_output_btn.clicked.connect(self.open_output_folder)
        button_layout.addWidget(self.start_btn)
        button_layout.addWidget(self.open_output_btn)
        left_layout.addLayout(button_layout)

        # Logs & Progress
        self.progress = QProgressBar()
        left_layout.addWidget(self.progress)
        left_layout.addWidget(QLabel("Process Engine Audit Logs:"))
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        left_layout.addWidget(self.log)

        # ==========================================
        # RIGHT HAND SIDE: DIRECTORY LIST & METADATA CARD
        # ==========================================
        right_panel = QVBoxLayout()
        right_panel.addWidget(QLabel("Target Directory Excel Files (.xlsx):"))
        
        self.file_list_widget = QListWidget()
        self.file_list_widget.setObjectName("styledFileList")
        self.file_list_widget.itemClicked.connect(self.request_metadata_load)
        
        # KEY CHANGES HERE: Hook up double click and custom context menus
        self.file_list_widget.itemDoubleClicked.connect(self.open_file_externally)
        self.file_list_widget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.file_list_widget.customContextMenuRequested.connect(self.show_right_click_menu)
        
        right_panel.addWidget(self.file_list_widget, stretch=3)

        right_panel.addWidget(QLabel("Workbook Quick Stats Summary:"))
        
        self.meta_card = QTextEdit()
        self.meta_card.setObjectName("metaSummaryCard")
        self.meta_card.setReadOnly(True)
        self.meta_card.setHtml("<p style='color: #64748b; font-style: italic;'>Select a spreadsheet file above to view structural details.</p>")
        right_panel.addWidget(self.meta_card, stretch=2)

        left_container = QWidget()
        left_container.setLayout(left_layout)
        right_container = QWidget()
        right_container.setLayout(right_panel)

        main_layout.addWidget(left_container, stretch=3)
        main_layout.addWidget(right_container, stretch=2)
        self.setLayout(main_layout)

        self.apply_ui_styles()
        self.toggle_task_inputs()

    # =====================================================================
    # NATIVE SYSTEM EXCEL FILE LAUNCHERS
    # =====================================================================
    def open_file_externally(self, item):
        """Launches the system's native default editor (e.g. Excel) for the chosen spreadsheet."""
        filename = item.text().replace("📊  ", "").strip()
        
        # Check against default messages inside the list block
        if filename.startswith("(") or filename.startswith("Reading Error"):
            return

        folder = self.input_path.text()
        full_file_path = os.path.normpath(os.path.join(folder, filename))

        if not os.path.exists(full_file_path):
            self.log_message(f"❌ Error: Path doesn't exist: {full_file_path}")
            return

        try:
            self.log_message(f"📂 Launching external system editor for: {filename}")
            if os.name == "nt":  # Windows
                os.startfile(full_file_path)
            elif sys.platform == "darwin":  # macOS
                subprocess.run(["open", full_file_path])
            else:  # Linux variants
                subprocess.run(["xdg-open", full_file_path])
        except Exception as e:
            self.log_message(f"❌ Failed to run external process application: {str(e)}")

    def show_right_click_menu(self, position):
        """Constructs and prompts a custom interactive right-click context menu."""
        item = self.file_list_widget.itemAt(position)
        if not item:
            return

        # Initialize the pop-up panel menu component
        context_menu = QMenu(self)
        
        # Match aesthetic colors with layout styles
        context_menu.setStyleSheet("""
            QMenu { background-color: #ffffff; border: 1px solid #cbd5e1; border-radius: 4px; padding: 4px; }
            QMenu::item { padding: 6px 20px; color: #1e293b; font-weight: 500; }
            QMenu::item:selected { background-color: #2563eb; color: #ffffff; border-radius: 3px; }
        """)
        
        open_action = context_menu.addAction("📂 Open in Excel")
        
        # Executed position coordinates calculated directly under the cursor location
        action = context_menu.exec(self.file_list_widget.mapToGlobal(position))
        if action == open_action:
            self.open_file_externally(item)

    # =====================================================================
    # METADATA PROCESSING LOGIC
    # =====================================================================
    def request_metadata_load(self, item):
        filename = item.text().replace("📊  ", "").strip()
        if filename.startswith("(") or filename.startswith("Reading Error"):
            return
            
        folder = self.input_path.text()
        full_file_path = os.path.normpath(os.path.join(folder, filename))

        if not os.path.exists(full_file_path):
            return

        self.meta_card.setHtml(f"<p style='color: #2563eb; font-weight: bold;'>⏳ Scanning structure metrics for {filename}...</p>")
        
        self.meta_thread = ExcelMetaWorker(full_file_path)
        self.meta_thread.meta_loaded_signal.connect(self.display_metadata_card)
        self.meta_thread.error_signal.connect(lambda err: self.meta_card.setHtml(f"<p style='color: #ef4444;'>❌ Error: {err}</p>"))
        self.meta_thread.start()

    def display_metadata_card(self, meta):
        html_content = f"""
        <table width="100%" cellpadding="4" style="font-family: 'Segoe UI', Arial; font-size: 13px; color: #1e293b;">
            <tr><td><b>📁 File Size:</b></td><td style="color: #2563eb; font-weight: 600;">{meta['file_size']}</td></tr>
            <tr><td><b>📅 Last Modified:</b></td><td>{meta['mod_date']}</td></tr>
            <tr><td><b>🔢 Total Sheets:</b></td><td>{meta['num_sheets']}</td></tr>
            <tr><td><b>📖 Sheet Names:</b></td><td style="color: #475569; font-style: italic;">{meta['sheet_names']}</td></tr>
            <tr style="background-color: #f8fafc;"><td colspan="2" style="border-top: 1px solid #e2e8f0; padding-top: 6px;"><b>📊 Active Sheet Diagnostics ({meta['active_sheet']}):</b></td></tr>
            <tr><td>&nbsp;&nbsp;&nbsp;&nbsp;• Row Count:</td><td style="font-weight: bold; color: #10b981;">{meta['row_count']} rows</td></tr>
            <tr><td>&nbsp;&nbsp;&nbsp;&nbsp;• Column Count:</td><td style="font-weight: bold; color: #f59e0b;">{meta['col_count']} columns</td></tr>
        </table>
        """
        self.meta_card.setHtml(html_content)

    # =====================================================================
    # GUI AUXILIARY INTERACTIONS
    # =====================================================================
    def populate_file_list(self):
        self.file_list_widget.clear()
        path = self.input_path.text()
        if os.path.exists(path) and os.path.isdir(path):
            try:
                files = [f for f in os.listdir(path) if f.lower().endswith('.xlsx')]
                if files:
                    for filename in files: self.file_list_widget.addItem(f"📊  {filename}")
                else:
                    self.file_list_widget.addItem("(No .xlsx spreadsheets found in path)")
            except Exception as e: self.file_list_widget.addItem(f"Reading Error: {str(e)}")

    def toggle_task_inputs(self):
        task_idx = self.task_selector.currentIndex()
        is_splitter = (task_idx == 0)
        self.equipment.setVisible(is_splitter)
        self.equipment_row_label.setVisible(is_splitter)
        self.chunk_size.setVisible(is_splitter)
        self.chunk_row_label.setVisible(is_splitter)
        is_aggregator = (task_idx == 2)

        self.template_widget.setVisible(is_aggregator)
        self.template_row_label.setVisible(is_aggregator)

        is_cleaner = (task_idx == 3)
        self.template_widget.setVisible(is_cleaner)
        self.template_row_label.setVisible(is_cleaner)
        self.equipment.setVisible(is_cleaner)
        self.equipment_row_label.setVisible(is_cleaner)


    def select_input_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Source Input Directory")
        if folder:
            self.input_path.setText(folder)
            task_idx = self.task_selector.currentIndex()
            if task_idx == 0: self.output_path.setText(os.path.join(folder, "split_data_files"))
            elif task_idx == 1: self.output_path.setText(os.path.join(folder, "renamed_files"))
            elif task_idx == 2: self.output_path.setText(os.path.join(folder, "collected_data"))
            elif task_idx == 3: self.output_path.setText(os.path.join(folder, "cleaned_data"))

    def select_output_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Output Target Directory")
        if folder: self.output_path.setText(folder)

    def select_template_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Base Template Excel File", "", "Excel Files (*.xlsx)")
        if file_path: self.template_path_field.setText(file_path)

    def log_message(self, msg):
        self.log.append(msg)

    def execute_current_task(self):
        task_idx = self.task_selector.currentIndex()
        input_folder = self.input_path.text()
        output_folder = self.output_path.text()
        self.progress.setValue(0)
        if not input_folder or not output_folder: return

        if task_idx == 0:
            chunk = self.chunk_size.text()
            try: chunk = int(chunk)
            except ValueError: return
            from splitter_engine import EquipmentProductivity
            self.engine = EquipmentProductivity(folder_path=input_folder, output_folder=output_folder, equipment=self.equipment.currentText().lower(), logger=self.log_message)
            self.engine.chunk_size = chunk
        elif task_idx == 1:
            from naming_engine import FileRenamingEngine
            self.engine = FileRenamingEngine(input_folder=input_folder, file_type=".xlsx")
        elif task_idx == 2:
            template_path = self.template_path_field.text()
            if not template_path or not os.path.exists(template_path): return
            from processing_engine import DataProcessingEngine
            self.engine = DataProcessingEngine(input_folder=input_folder, template_path=template_path)

        elif task_idx == 3:
            template_path = self.template_path_field.text()
            if not template_path or not os.path.exists(template_path): return
            from cleaning_engine import DataCleaningEngine
            self.engine = DataCleaningEngine(
                {
                    'input_folder':input_folder, 
                    'output_folder':output_folder, 
                    'template_path':template_path, 
                    'logger':self.log_message, 
                    'progress_callback':self.progress.setValue
                }
            )
                


        self.worker = Worker(self.engine)
        self.worker.log_signal.connect(self.log_message)
        self.worker.progress_signal.connect(self.progress.setValue)
        self.worker.finished_signal.connect(lambda: [self.log_message("✅ Operational Run Finished Complete!"), self.populate_file_list()])
        self.worker.start()

    def open_output_folder(self):
        folder = self.output_path.text()
        if not folder or not os.path.exists(folder): return
        if os.name == "nt": os.startfile(folder)
        elif sys.platform == "darwin": subprocess.run(["open", folder])
        else: subprocess.run(["xdg-open", folder])

    def apply_ui_styles(self):
        self.setStyleSheet("""
            QWidget { 
                font-size: 13px; 
                background-color: #f1f5f9; 
                color: #334155;
            }
            QLineEdit, QComboBox {
                border: 1px solid #cbd5e1; 
                border-radius: 6px;
                padding: 6px;
                background: white;
                font-size: 14px;
                color: #1e293b;
            }
            QLineEdit:focus, QComboBox:focus, QListWidget:focus { border: 2px solid #2563eb; }
            QTextEdit { border: 1px solid #cbd5e1; border-radius: 6px; background: #1e293b; color: #38bdf8; font-family: 'Consolas', monospace; }
            QPushButton {
                background-color: #2563eb; color: white; border: none;
                border-radius: 6px; padding: 8px 14px; font-weight: 600; min-height: 32px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
            QLabel { font-size: 14px; font-weight: bold; color: #334155; margin-top: 4px; }
            QProgressBar { border: 1px solid #cbd5e1; border-radius: 4px; text-align: center; background: #e2e8f0; font-weight: bold; }
            QProgressBar::chunk { background-color: #2563eb; }
            
            /* Directory File List View Styles */
            QListWidget#styledFileList {
                background-color: white; border: 1px solid #cbd5e1;
                border-radius: 8px; padding: 5px; outline: none;
            }
            QListWidget#styledFileList::item {
                padding: 10px 12px;
                margin: 2px 4px; 
                border-radius: 6px;
                background-color: #f8fafc;
                border: 1px solid #e2e8f0;
                color: #1e293b; 
                font-weight: 500; 
                font-family: 'Segoe UI', Arial;
            }
            QListWidget#styledFileList::item:hover {
                background-color: #eff6ff; 
                           color: #2563eb; 
                           border: 1px solid #bfdbfe;
            }
            QListWidget#styledFileList::item:selected {
                background-color: #2563eb; 
                color: white; 
                font-weight: bold; 
                border: 1px solid #2563eb;
            }

            /* Diagnostic Metadata Panel Card Styling */
            QTextEdit#metaSummaryCard {
                background-color: #ffffff;
                border: 1px solid #cbd5e1;
                border-radius: 8px;
                padding: 8px;
            }
        """)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = CMIDataProcessorGUI()
    window.show()
    sys.exit(app.exec())