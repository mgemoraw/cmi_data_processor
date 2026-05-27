"""
Data Cleaning Engine
"""
import os
import sys
from openpyxl import load_workbook, Workbook
from pathlib import Path 
import re 
import shutil
import random
import sys

sys.stdout.reconfigure(encoding='utf-8')

class DataCleaningEngine:
    SHEET_ALIASES = {
            "daily_variables": [
                "daily_variables",
                "daily variable",
                "daily variables",
                "daily-variables",
            ],

            "mpdm": [
                "mpdm",
                "MPDM"
            ],
            "work_sampling": [
                "Work Sampling",
                "work_sampling",
                "work sampling",
            ],
            "equipment": [
                "Labor",
                "Truck",
                "Dozer",
                "Excavator",
                "Roller",
                "Grader",
            ]
        }
    
    def __init__(self, input_folder, template_path, output_path=None, logger=None,  equipment=None, ):

        self.source_path = Path(input_folder) if input_folder else Path.cwd()
        # self.output_path = Path(self.output_folder) if self.output_path else Path.cwd()

        self.template_path = Path(template_path) if template_path else Path.cwd()
        self.template_mapping = {}

        self.logger = logger if logger else print 

        self.progress_callback = None
        
        self.output_folder = os.path.join(self.source_path, "cleaned_files")

        # Pattern for DD-MM-YYYY
        # self.date_pattern = r"(\d{2}-\d{2}-\d{4})"
        self.date_pattern = r"(\d{2}[-_/]\d{2}[-_/]\d{4})"
        self.data_count_pattern  = r"(\d{2}[-_/]\d{2}[-_/]\d{2}[-_/]\d{4})"
        self.equipment = equipment

    def log(self, message):
        self.logger(message)


    def update_progress(self, value):
        if self.progress_callback:
            self.progress_callback(value)
    
    def start_cleaning(self):
        for index, file in enumerate(os.listdir(self.source_path)):
            if file.endswith('.xlsx') and file != self.template_path.name:
                self.logger(f"📂 Processing file: {file}")
                self._process_cleaning(file)
                self._copy_template(template_path=self.template_path)

 


    def _process_cleaning(self, file):
        source_wb = load_workbook(os.path.join(self.source_path, file), read_only=True)
        sheetnames = source_wb.sheetnames
        self.logger(f"Sheets names found: {sheetnames}")

        if self.equipment is not None:
            self.equipment = self.equipment.lower()
        

        new_wb = self._copy_template(template_path=self.template_path)

        if self.equipment in sheetnames:
            equipment_ws = source_wb[self.equipment]

            self.equipment_handlers = {
                "dozer": self._clean_dozer_records,
                "excavator": self._clean_excavator_records,
                "truck": self._clean_truck_records,
                "roller": self._clean_roller_records,
            }

            handler = self.equipment_handlers.get(self.equipment)

            if handler:
                handler(equipment_ws)
            
        
        self._clean_mpdm_records(new_wb)

        self._clean_daily_variables(new_wb)

        

    def _clean_equipment_records(self, worksheet):
        pass 
    
    def _clean_dozer_records(self, worksheet):
        self._clean_blanks(worksheet) 

    def _clean_excavator_records(self, worksheet):
        pass 

    def _clean_truck_records(self, worksheet):
        pass 
    def _clean_labor_records(self, worksheet):
        pass
    def _clean_work_sampling_records(self, worksheet):
        pass 

    def _clean_mpdm_records(self, worksheet):
        pass 

    def _clean_daily_variables(self, worksheet):
        pass 
    
    def _clean_roller_records(self, worksheet):
        pass 

    def _clean_grader_records(self, worksheet):
        pass 

    def _copy_template(self, template_path, file_name=None, counter=0):
        """
        Copy template workbook into output folder
        with a new filename.
        """
        template_source = Path(template_path)

        if file_name is None:
            # -----------------------------------
            # Generate random filename if missing
            # -----------------------------------
            if file_name is None:

                file_name = (f"{random.randint(0,9999)}")
        
        # -----------------------------------
        # Extract template filename stem
        # template.xlsx -> template
        # -----------------------------------
        template_stem = template_source.stem

        # -----------------------------------
        # Extract suffix
        # .xlsx
        # -----------------------------------
        template_suffix = template_source.suffix

        # -----------------------------------
        # Create new filename
        # -----------------------------------
        new_filename = (
            f"{file_name}_"
            f"{template_suffix}"
        )

        # -----------------------------------
        # Destination path
        # -----------------------------------
        destination = (
            Path(self.output_folder)
            / new_filename
        )

        # Ensure output folder exists
        destination.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        # -----------------------------------
        # Copy template
        # -----------------------------------
        shutil.copy2(
            template_source,
            destination
        )

        self.logger(
            f"✅ Template copied: {new_filename}"
        )

        return destination



    def read_excel_contents(self):
        self.logger("... Started data cleaning...")
        # Implement the main processing logic here
        self.logger(f"📂 Source folder: {self.source_path}")
        # template = self.read_template_file(self.template_path)
        # files = self.sort_files_by_date(os.listdir(self.source_path))

        for index, file in enumerate(os.listdir(self.source_path)):
            self.logger(f"... Processing file: {file}")

    def _clean_blanks(self, ws):
        # ws = workbook['dozer']
        date = ws['A7'].value
        project_code = ws['B7'].value
        data_collector = ws['C7'].value
        number_of_equipment_types = ws['D7'].value

        self.logger(f"Date: {date}\n Project Code: {project_code}\n DataCollector: {data_collector}\n Number of Equipment Types: {number_of_equipment_types}")





if __name__ == "__main__":
    cleaner = DataCleaningEngine(
        input_folder="input_folder/", 
        template_path="forms/Dozer-Template.xlsx", 
        equipment="Dozer",
    )

    cleaner.start_cleaning()
