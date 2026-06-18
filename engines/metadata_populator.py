from openpyxl import Workbook, load_workbook
from pathlib import Path 
from dataclasses import dataclass
from datetime import datetime
import os 
import re 




# =====================================================================
# DATA AGGREGATING AND ANALYSIS ENGINE COMPONENTS
# =====================================================================
@dataclass
class Metadata:
    date: str
    project_code: str
    operation: str
    equipment: str  # equipment type
    data_count: int
    equipment_types: int # number of equipment types
    data_collector: str # data collector name

    def to_dict(self):
        return Metadata.__dict__()
    

class MetadataPopulator:
    def __init__(self, source_folder, equipment, logger = print, progress_callback=None):
        self.source_folder = source_folder
        self.equipment = equipment
        self.date_pattern = r"(\d{2}[-_/]\d{2}[-_/]\d{4})"

        self.logger = logger 
        self.logger = logger
        self.progress_callback = progress_callback

    def log(self):
        return self.logger

    def _log(self, text):
        if self.logger:
            self.logger(text)
        else:
            print(text)

    def _set_progress(self, val):
        if self.progress_callback:
            self.progress_callback(val)

    def read_excel_contents(self):
        """Standard execution hook matching your Worker class thread setup."""
        self._process_files()

    def _get_date_from_filename(self, file_path):
        # self.logger(f"reading date: from {file_path}")
        try:
            
            file_name = os.path.basename(file_path)
            mach = re.search(self.date_pattern, file_name)
            if mach:
                return mach.group(1)
            print(mach)
        except Exception as e:
            return None

    def _parse_date_str(self, string:str):
        
        mach = re.search(self.date_pattern, string)
        if mach:
            return mach.group(1)
        else:
            return None 
        
    def _get_date_str(self, file_path):
        self.logger(f"reading date: from {file_path}")
        try:
            
            file_name = os.path.basename(file_path)
            mach = re.search(self.date_pattern, file_name)
            if mach:
                return mach.group(1)


            wb = load_workbook(file_path, data_only=True)
            ws = wb['daily_variables']
            date_value = ws['H5'].value
            if hasattr(date_value, "strftime"):
                return date_value.strftime("%d-%m-%Y")
            
            else:
                if "/" in str(date_value):
                    return date_value.replace("/", "-")
                if "_" in str(date_value):
                    return date_value.replace("_", "-")
                return date_value

        except Exception as e:
            self.logger(f"❌ Error reading date from {file_path}: {e}")
            return None

    # def run(self):
    #     self._process_files()

    def start(self):
        self._process_files()


    def _normalize_date_str(self, value):
        date_ = self._parse_date_str(value.strip())
        if "/" in str(value):
            return date_.replace("/", "-")
        if "_" in str(value):
            return date_.replace("_", "-")
        
        return date_

    def sort_files_by_date(self, files):
        def extract_date_object(file_name):
            match = re.search(self.date_pattern, file_name)

            if match:
                date_str = match.group(1)  # Extracts the date substring (e.g., "02-03-2014")
                try:
                    # Convert the string to a real datetime object for accurate chronological sorting
                    # Adjust "%d-%m-%Y" if your pattern captures dates differently (e.g., "%Y-%m-%d")
                    return datetime.strptime(date_str, "%d-%m-%Y")
                except ValueError:
                    # Handle cases where the regex matches a string that isn't a valid calendar date
                    return datetime.max
            
            # If no date pattern is found, return datetime.max to safely push the file to the end of the list
            return datetime.max

        # Sort using the actual datetime object instead of a string
        return sorted(files, key=extract_date_object)

    def list_excel_files(self):
        """
        Return all Excel files in the folder
        """
        return list(self.source_folder.glob("*.xlsx"))
        
    def _process_files(self):
        files = self.sort_files_by_date(os.listdir(self.source_folder))
        # files = os.listdir(self.source_folder)
        total_files = self.list_excel_files()

        metadata = { }
            
    
        for index, file in enumerate(files):
            if file.endswith('.xlsx'):
                date_str = self._get_date_from_filename(file)
                # date_str = self._normalize_date_str(date_str)

                # metadata[date_str] =  {
                #     'date': None,
                #     'project_code': None,
                #     'operation': None,
                #     'equipment_types': None,
                #     'data_collector': None,
                #     'equipment': self.equipment,
                # }
               
                if not date_str:
                    print(f"{file} -- date is is None")

                source_path = os.path.join(self.source_folder, file)
                
                # print(files)
                try:
                    if  file.startswith(date_str) and file.endswith("part_1.xlsx"):
                        wb = load_workbook(source_path, data_only=True)
                        ws = wb[self.equipment.lower()]
                        metadata[date_str] = {
                            "date": self._normalize_date_str(ws['A7'].value),
                            "project_code": ws['B7'].value,
                            "equipment_types": ws['D7'].value,
                            "data_collector": ws['C7'].value,
                            "operation" : wb['mpdm']['E7'].value
                        }

                        self.logger(f"catched metadata from {file}")
                        # self.logger(f"Metadata: {metadata.get(date_str)}")
                        
                    
                    elif (file.startswith(date_str)) and not (file.endswith('_part_1.xlsx')):
                        self.logger(f"updating metadata on  {file}")
                        wb = load_workbook(source_path)
                        ws = wb[self.equipment.lower()]
                        ws['A7'] = metadata.get(date_str)['date']
                        ws['B7'] = metadata.get(date_str)['project_code']
                        ws['C7'] = metadata.get(date_str)['data_collector']
                        ws['D7'] = metadata.get(date_str)['equipment_types']

                        mpdm_ws = wb['mpdm']
                        mpdm_ws['A7'] = metadata.get(date_str)['date']
                        mpdm_ws['B7'] = metadata.get(date_str)['project_code']
                        mpdm_ws['C7'] = metadata.get(date_str)['data_collector']

                        print(f"Metadata of file: {metadata.get(date_str)}")
                        
                        wb.save(source_path)
                        wb.close()

                        self.logger(f"Metadata inserted to {file} successfully")
                    else:
                        self.logger(f"{date_str} in {file} and {metadata.get(date_str)['date']} do not match ")

                except Exception as e:
                    raise e

if __name__ == "__main__":

    # source_folder = Path("G:\CMI-NPN\CMI-DATABASE\TEMPORARY\P013-excavator-split_data_files_80").resolve()
    # source_folder = Path("G:\CMI-NPN\CMI-DATABASE\TEMPORARY\P005-excavator-split_data_files_80").resolve()
    source_folder = Path('G:\CMI-NPN\CMI_DATA_PROCESSOR\P012-excavator-split_data_files').resolve()
    equipment = "excavator"

    mp = MetadataPopulator(
        source_folder=source_folder,
        equipment=equipment
    )
    mp.start()


   
