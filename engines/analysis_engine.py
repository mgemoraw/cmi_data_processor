import sys
import os
import subprocess
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass
from openpyxl import load_workbook, Workbook
import win32com.client as win32



# =====================================================================
# DATA AGGREGATING AND ANALYSIS ENGINE COMPONENTS
# =====================================================================
@dataclass
class OutputData:
    day: str
    date: str
    operation_activity: str
    equipment_type: str
    description_work: str
    productivity: float
    ideal_productivity: float
    overall_method_productivity: float
    ideal_cycle_variability: float
    overall_cycle_variability: float
    data_count: str = None

    HEADERS = [
        "Day", 
        "Date", 
        "Operation/Activity", 
        "Equipment Type", 
        "Description Work",
        "Productivity (Units/hr)", 
        "Ideal Productivity (Unit/hr)", 
        "Overall Method Productivity (Unit/hr)", 
        "Ideal Cycle Variability (%)", 
        "Overall Cycle Variability (%)",
        "Data Source File"
    ]

    @classmethod
    def write_headers(cls, ws, row=1):
        for col, header in enumerate(cls.HEADERS, start=1):
            ws.cell(row=row, column=col, value=header)

    def write_row(self, ws, row, data_count=None):
        values = [
            self.day, 
            self.date, 
            self.operation_activity, 
            self.equipment_type, 
            self.description_work,
            self.productivity, 
            self.ideal_productivity, 
            self.overall_method_productivity,
            self.ideal_cycle_variability, 
            self.overall_cycle_variability,
            data_count,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value)


class AnalysisEngine:
    def __init__(self, data_folder, template_path=None, equipment=None, particular=None, activity=None, logger=None, progress_callback=None):
        self.source_folder = Path(data_folder)
        self.output_folder = self.source_folder 
        self.template_path = Path(template_path) if template_path else None
        self.equipment = equipment
        self.particular = particular if particular else f"{equipment} Work Description"
        self.activity = activity if activity else f"{equipment} Activity Process"
        
        self.logger = logger
        self.progress_callback = progress_callback
        self.output_folder.mkdir(exist_ok=True)

    def _log(self, text):
        if self.logger:
            self.logger(text)
        else:
            print(text)

    def _set_progress(self, val):
        if self.progress_callback:
            self.progress_callback(val)

    def read_excel_contents(self):
        """Standard execution hook matching your Worker class thread setup."""
        self.run()

    def refresh_excel_file(self, file_path):
        try:
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(str(file_path))
            workbook.RefreshAll()
            workbook.Save()
            workbook.Close()
            excel.Quit()
        except Exception as e:
            self._log(f"⚠️ COM Connection Notice (Refresh): {e}")

    def run(self):
        self._log("🏁 Starting Analysis Engine Aggregation...")
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = self.activity[:30] # Excel limit tab rule boundary restriction max length 30

        OutputData.write_headers(new_ws, row=1)
        current_row = 2

        all_files = [f for f in os.listdir(self.source_folder) if f.endswith(".xlsx")]
        total_files = len(all_files)

        if total_files == 0:
            self._log("⚠️ No active spreadsheet data logs located inside target input folder directory.")
            return False

        for index, file in enumerate(all_files):
            full_path = os.path.join(self.source_folder, file)
            self._log(f"🔄 Re-evaluating Excel connections for: {file}")
            self.refresh_excel_file(full_path)

            try:
                wb = load_workbook(filename=full_path, data_only=True)
                if self.equipment.title() in wb.sheetnames:
                    sheet_name = self.equipment.title()
                else:
                    sheet_name = f'{self.equipment.title()}1'

                if sheet_name not in wb.sheetnames:
                    self._log(f"⚠️ Target sheet layout segment '{sheet_name}' missing inside workbook {file}. Skipping.")
                    continue
                    
                ws = wb[sheet_name]
                
                # source target data  row
                start_row = 11

                pr_col = self.get_target_cells()['productivity']
                ideal_col = self.get_target_cells()['ideal_productivity']
                overall_col = self.get_target_cells()['overall_method_productivity']
                ideal_cycle_col = self.get_target_cells()['ideal_cycle_variability']
                overall_cycle_col = self.get_target_cells()['overall_cycle_variability']

                data = OutputData(
                    day=str(index + 1),
                    date=self.extract_date(ws),
                    operation_activity=self.activity,
                    equipment_type=self.equipment,
                    description_work=self.particular,
                    productivity=self.extract_average_productivity(wb, ws, pr_col),
                    ideal_productivity=ws[f'{ideal_col}{start_row}'].value,
                    overall_method_productivity=ws[f'{overall_col}{start_row}'].value,
                    ideal_cycle_variability=ws[f"{ideal_cycle_col}{start_row}"].value,
                    overall_cycle_variability=ws[f'{overall_cycle_col}{start_row}'].value,
                )
                data_count=self.extract_data_count(file)
                data.write_row(new_ws, current_row, data_count=data_count)
                current_row += 1
                self._log(f"✅ Extracted Row Stats successfully from: {file}")
            except Exception as ex:
                self._log(f"❌ Structural extraction exception on file sequence: {file}. details: {ex}")
            
            # Update Progress Bar UI Component safely
            self._set_progress(int(((index + 1) / total_files) * 100))

        output_file = os.path.join(self.output_folder, "Equipment_Productivity.xlsx")
        new_wb.save(output_file)
        self._log(f"💾 Aggregated Summary Report saved successfully to directory file location: {output_file}")
        return True
    
    def get_target_cells(self):
        if self.equipment.lower() == 'dozer':
            return {
                'productivity': "N",
                'ideal_productivity': "O",
                'overall_method_productivity' :"P",
                'ideal_cycle_variability': "Q",
                'overall_cycle_variability': "R",
            }
        elif self.equipment.lower() == 'excavator':
            return {
                'productivity': "O",
                'ideal_productivity': "P",
                'overall_method_productivity' :"Q",
                'ideal_cycle_variability': "R",
                'overall_cycle_variability': "S",
            }
        else:
            return {}


    def extract_data_count(self, file):
        name = Path(file).stem

        return name
    
    def extract_date(self, ws):
        l6 = ws["L6"].value 
        m6 = ws["M6"].value
        n6 = ws["N6"].value
        o6 = ws["O6"].value

        # Case 1: Split entries (Day in L6, Month in M6, Year in N6)
        if (m6 is not None and n6 is not None) and self.equipment.lower() == 'dozer':
            # e.g., "05/12/2026" or "5-December-2026"
            return f"{l6}/{m6}/{n6}"
        
        elif (n6 is not None and o6 is not None) and self.equipment.lower() in ['excavator', 'truck']:
            # e.g., "05/12/2026" or "5-December-2026"
            return f"{m6}/{n6}/{o6}"
        
        # Case 2: Merged cells (L6 holds the full date value, M6 & N6 are None)
        elif l6 is not None and self.equipment.lower() == 'dozer':
            # If Excel already parsed it as a datetime object, format it nicely
            if hasattr(l6, "strftime"):
                return l6.strftime("%Y-%m-%d")
            return str(l6)
        
        elif m6 is not None and self.equipment.lower() in ['excavator', 'truck']:
            # If Excel already parsed it as a datetime object, format it nicely
            if hasattr(m6, "strftime"):
                return m6.strftime("%Y-%m-%d")
            return str(m6)
        
        # Case 3: Completely empty fallback
        return "N/A"
    
    def update_template_formulas(self, file):
        pass

    def extract_average_productivity(self, source_wb, source_ws, source_column='N'):
        # file = Path(file)
        # wb = load_workbook(file)
        # equipment_ws = None
        # for sheetname in  wb.sheetnames:
        #     if sheetname.lower().startswith(self.equipment.lower()):
        #         equipment_ws = wb[sheetname]
        #         break
        # equipment_ws['N111']

        # if  source_ws not in source_wb.sheetnames:
        #     self._log(f"❌ Sheet '{source_ws}' not found in {source_wb}")
        #     source_wb.close()
        #     return 0
        
        data = []
        for row in range(11, 112):
            cell_value = source_ws[f'{source_column}{row}'].value
            # if cell_value is not None:
            try:
                data.append(float(cell_value))
            except (ValueError, TypeError):
                # Ignore text rows
                continue
        source_wb.close()  # free up system memory early


        # calculate the average safely
        if not data:
            print("⚠️ No numeric values found in range N11:N111. Average cannot be computed.")
            calculated_average = "N/A"
        else:
            calculated_average = sum(data) / len(data)

        return calculated_average


