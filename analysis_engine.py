import sys
import os
import subprocess
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass
from PySide6.QtWidgets import (
    QFormLayout, QApplication, QWidget, QVBoxLayout,
    QPushButton, QFileDialog, QLineEdit, QLabel, 
    QTextEdit, QHBoxLayout, QComboBox, QProgressBar, QListWidget,
    QMenu
)
from PySide6.QtCore import QThread, Signal, Qt
from openpyxl import load_workbook, Workbook
import win32com.client as win32

# =====================================================================
# THREAD WORKER - ASYNCHRONOUS METADATA EXTRACTOR
# =====================================================================
class ExcelMetaWorker(QThread):
    meta_loaded_signal = Signal(dict)
    error_signal = Signal(str)

    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path

    def run(self):
        try:
            stat_info = os.stat(self.file_path)
            file_size_kb = round(stat_info.st_size / 1024, 2)
            mod_time = datetime.fromtimestamp(stat_info.st_mtime).strftime('%Y-%m-%d %H:%M:%S')

            wb = load_workbook(self.file_path, read_only=True)
            sheet_names = wb.sheetnames
            num_sheets = len(sheet_names)
            
            ws = wb.active
            active_sheet_title = ws.title
            max_rows = ws.max_row if ws.max_row else "Unknown (Streaming Required)"
            max_cols = ws.max_column if ws.max_column else "Unknown"

            wb.close()

            metadata = {
                "file_size": f"{file_size_kb} KB",
                "mod_date": mod_time,
                "num_sheets": str(num_sheets),
                "sheet_names": ", ".join(sheet_names),
                "active_sheet": active_sheet_title,
                "row_count": str(max_rows),
                "col_count": str(max_cols)
            }
            self.meta_loaded_signal.emit(metadata)
        except Exception as e:
            self.error_signal.emit(str(e))


# =====================================================================
# THREAD WORKER - CORE ENGINES
# =====================================================================
class Worker(QThread):
    log_signal = Signal(str)
    progress_signal = Signal(int)
    finished_signal = Signal()

    def __init__(self, engine):
        super().__init__()
        self.engine = engine

    def run(self):
        # Attach dynamic logging/progress emitters to the core engine object
        self.engine.logger = self.log_signal.emit
        self.engine.progress_callback = self.progress_signal.emit
        
        if hasattr(self.engine, 'read_excel_contents'):
            self.engine.read_excel_contents()
        self.finished_signal.emit()


# =====================================================================
# DATA AGGREGATING AND ANALYSIS ENGINE COMPONENTS
# =====================================================================
@dataclass
class OutputData:
    day: str
    date: str
    operation_activity: str
    equipment_type: str
    description_work: str
    productivity: float
    ideal_productivity: float
    overall_method_productivity: float
    ideal_cycle_variability: float
    overall_cycle_variability: float

    HEADERS = [
        "Day", "Date", "Operation/Activity", "Equipment Type", "Description Work",
        "Productivity (Units/hr)", "Ideal Productivity (Unit/hr)", 
        "Overall Method Productivity (Unit/hr)", "Ideal Cycle Variability (%)", "Overall Cycle Variability (%)"
    ]

    @classmethod
    def write_headers(cls, ws, row=1):
        for col, header in enumerate(cls.HEADERS, start=1):
            ws.cell(row=row, column=col, value=header)

    def write_row(self, ws, row):
        values = [
            self.day, 
            self.date, 
            self.operation_activity, 
            self.equipment_type, 
            self.description_work,
            self.productivity, 
            self.ideal_productivity, 
            self.overall_method_productivity,
            self.ideal_cycle_variability, 
            self.overall_cycle_variability
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value)


class AnalysisEngine:
    def __init__(self, data_folder, template_path=None, equipment=None, particular=None, activity=None, logger=None, progress_callback=None):
        self.source_folder = Path(data_folder)
        self.output_folder = self.source_folder / "output"
        self.template_path = Path(template_path) if template_path else None
        self.equipment = equipment
        self.particular = particular if particular else f"{equipment} Work Description"
        self.activity = activity if activity else f"{equipment} Activity Process"
        
        self.logger = logger
        self.progress_callback = progress_callback
        self.output_folder.mkdir(exist_ok=True)

    def _log(self, text):
        if self.logger:
            self.logger(text)
        else:
            print(text)

    def _set_progress(self, val):
        if self.progress_callback:
            self.progress_callback(val)

    def read_excel_contents(self):
        """Standard execution hook matching your Worker class thread setup."""
        self.run()

    def refresh_excel_file(self, file_path):
        try:
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(str(file_path))
            workbook.RefreshAll()
            workbook.Save()
            workbook.Close()
            excel.Quit()
        except Exception as e:
            self._log(f"⚠️ COM Connection Notice (Refresh): {e}")

    def run(self):
        self._log("🏁 Starting Analysis Engine Aggregation...")
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = self.activity[:30] # Excel limit tab rule boundary restriction max length 30

        OutputData.write_headers(new_ws, row=1)
        current_row = 2

        all_files = [f for f in os.listdir(self.source_folder) if f.endswith(".xlsx")]
        total_files = len(all_files)

        if total_files == 0:
            self._log("⚠️ No active spreadsheet data logs located inside target input folder directory.")
            return False

        for index, file in enumerate(all_files):
            full_path = os.path.join(self.source_folder, file)
            self._log(f"🔄 Re-evaluating Excel connections for: {file}")
            self.refresh_excel_file(full_path)

            try:
                wb = load_workbook(filename=full_path, data_only=True)
                if self.equipment.title() in wb.sheetnames:
                    sheet_name = self.equipment.title()
                else:
                    sheet_name = f'{self.equipment.title()}1'

                if sheet_name not in wb.sheetnames:
                    self._log(f"⚠️ Target sheet layout segment '{sheet_name}' missing inside workbook {file}. Skipping.")
                    continue
                    
                ws = wb[sheet_name]

                data = OutputData(
                    day=str(index + 1),
                    date=self.extract_date(ws),
                    operation_activity=self.activity,
                    equipment_type=self.equipment,
                    description_work=self.particular,
                    productivity=ws["N11"].value, 
                    ideal_productivity=ws["O11"].value,
                    overall_method_productivity=ws["P11"].value,
                    ideal_cycle_variability=ws["Q11"].value,
                    overall_cycle_variability=ws["R11"].value,
                )

                data.write_row(new_ws, current_row)
                current_row += 1
                self._log(f"✅ Extracted Row Stats successfully from: {file}")
            except Exception as ex:
                self._log(f"❌ Structural extraction exception on file sequence: {file}. details: {ex}")
            
            # Update Progress Bar UI Component safely
            self._set_progress(int(((index + 1) / total_files) * 100))

        output_file = os.path.join(self.output_folder, f"{self.activity}.xlsx")
        new_wb.save(output_file)
        self._log(f"💾 Aggregated Summary Report saved successfully to directory file location: {output_file}")
        return True
    
    def extract_date(self, ws):
        l6 = ws["L6"].value 
        m6 = ws["M6"].value
        n6 = ws["N6"].value

        # Case 1: Split entries (Day in L6, Month in M6, Year in N6)
        if m6 is not None and n6 is not None:
            # e.g., "05/12/2026" or "5-December-2026"
            return f"{l6}/{m6}/{n6}"
        
        # Case 2: Merged cells (L6 holds the full date value, M6 & N6 are None)
        elif l6 is not None:
            # If Excel already parsed it as a datetime object, format it nicely
            if hasattr(l6, "strftime"):
                return l6.strftime("%Y-%m-%d")
            return str(l6)
        
        # Case 3: Completely empty fallback
        return "N/A"


# =====================================================================
# MAIN WINDOW WITH INTEGRATED ANALYSIS ENGINE
# =====================================================================
class CMIDataProcessorGUI(QWidget):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("CMI-Data Processor (Multi-Task Pro)")
        self.setMinimumSize(950, 700)
        self.init_ui()

    def init_ui(self):
        main_layout = QHBoxLayout()
        left_layout = QVBoxLayout()
        form = QFormLayout()

        # Task Selector Updated
        self.task_selector = QComboBox()
        self.task_selector.addItems([
            "Excel File Splitting",
            "File Renaming (Date Match/G4 Extraction)",
            "Data Aggregation & Template Counter",
            "Clean Data",
            "Data Analysis Engine" 
        ])
        self.task_selector.currentIndexChanged.connect(self.toggle_task_inputs)
        form.addRow("Select Action Task:", self.task_selector)

        # Directory Fields
        input_widget = QWidget()
        input_layout = QHBoxLayout(input_widget)
        self.input_path = QLineEdit()
        self.input_path.textChanged.connect(self.populate_file_list)
        btn_input = QPushButton("Browse")
        btn_input.clicked.connect(self.select_input_folder)
        input_layout.addWidget(self.input_path)
        input_layout.addWidget(btn_input)
        form.addRow("Input Folder:", input_widget)

        output_widget = QWidget()
        output_layout = QHBoxLayout(output_widget)
        self.output_path = QLineEdit()
        btn_output = QPushButton("Browse")
        btn_output.clicked.connect(self.select_output_folder)
        output_layout.addWidget(self.output_path)
        output_layout.addWidget(btn_output)
        form.addRow("Output Folder:", output_widget)

        # Dynamic Parameters
        self.equipment = QComboBox()
        self.equipment.addItems(["Excavator", "Dozer", "Loader", "Grader", "Roller", "Truck", "Labor"])
        self.equipment_row_label = QLabel("Equipment Type:")
        form.addRow(self.equipment_row_label, self.equipment)

        self.chunk_size = QLineEdit()
        self.chunk_size.setPlaceholderText("100")
        self.chunk_row_label = QLabel("Chunk Size:")
        form.addRow(self.chunk_row_label, self.chunk_size)

        self.template_path_field = QLineEdit()
        self.template_path_field.setPlaceholderText("Path to baseline template .xlsx file...")
        self.template_btn = QPushButton("Browse Template")
        self.template_btn.clicked.connect(self.select_template_file)
        
        self.template_widget = QWidget()
        t_lay = QHBoxLayout(self.template_widget)
        t_lay.addWidget(self.template_path_field)
        t_lay.addWidget(self.template_btn)
        t_lay.setContentsMargins(0, 0, 0, 0)
        
        self.template_row_label = QLabel("Template File:")
        form.addRow(self.template_row_label, self.template_widget)

        # Particular and Activity Inputs (Added for Data Analysis Engine)
        self.particular_field = QLineEdit()
        self.particular_field.setPlaceholderText("e.g., Site Clearing using Dozer")
        self.particular_row_label = QLabel("Particular:")
        form.addRow(self.particular_row_label, self.particular_field)

        self.activity_field = QLineEdit()
        self.activity_field.setPlaceholderText("e.g., Site clearing operations")
        self.activity_row_label = QLabel("Activity:")
        form.addRow(self.activity_row_label, self.activity_field)

        left_layout.addLayout(form)

        # Action Buttons
        button_layout = QHBoxLayout()
        self.start_btn = QPushButton("🚀 Start Process")
        self.start_btn.clicked.connect(self.execute_current_task)
        self.start_btn.setStyleSheet("background-color: #10b981; min-height: 38px;")
        self.open_output_btn = QPushButton("📂 Open Output Location")
        self.open_output_btn.clicked.connect(self.open_output_folder)
        button_layout.addWidget(self.start_btn)
        button_layout.addWidget(self.open_output_btn)
        left_layout.addLayout(button_layout)

        # Logs & Progress
        self.progress = QProgressBar()
        left_layout.addWidget(self.progress)
        left_layout.addWidget(QLabel("Process Engine Audit Logs:"))
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        left_layout.addWidget(self.log)

        # RIGHT HAND SIDE
        right_panel = QVBoxLayout()
        right_panel.addWidget(QLabel("Target Directory Excel Files (.xlsx):"))
        
        self.file_list_widget = QListWidget()
        self.file_list_widget.setObjectName("styledFileList")
        self.file_list_widget.itemClicked.connect(self.request_metadata_load)
        self.file_list_widget.itemDoubleClicked.connect(self.open_file_externally)
        self.file_list_widget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.file_list_widget.customContextMenuRequested.connect(self.show_right_click_menu)
        
        right_panel.addWidget(self.file_list_widget, stretch=3)
        right_panel.addWidget(QLabel("Workbook Quick Stats Summary:"))
        
        self.meta_card = QTextEdit()
        self.meta_card.setObjectName("metaSummaryCard")
        self.meta_card.setReadOnly(True)
        self.meta_card.setHtml("<p style='color: #64748b; font-style: italic;'>Select a spreadsheet file above to view structural details.</p>")
        right_panel.addWidget(self.meta_card, stretch=2)

        left_container = QWidget()
        left_container.setLayout(left_layout)
        right_container = QWidget()
        right_container.setLayout(right_panel)

        main_layout.addWidget(left_container, stretch=3)
        main_layout.addWidget(right_container, stretch=2)
        self.setLayout(main_layout)

        self.apply_ui_styles()
        self.toggle_task_inputs()

    def open_file_externally(self, item):
        filename = item.text().replace("📊  ", "").strip()
        if filename.startswith("(") or filename.startswith("Reading Error"):
            return
        folder = self.input_path.text()
        full_file_path = os.path.normpath(os.path.join(folder, filename))
        if not os.path.exists(full_file_path):
            self.log_message(f"❌ Error: Path doesn't exist: {full_file_path}")
            return
        try:
            self.log_message(f"📂 Launching external system editor for: {filename}")
            if os.name == "nt":
                os.startfile(full_file_path)
            elif sys.platform == "darwin":
                subprocess.run(["open", full_file_path])
            else:
                subprocess.run(["xdg-open", full_file_path])
        except Exception as e:
            self.log_message(f"❌ Failed to run external process application: {str(e)}")

    def show_right_click_menu(self, position):
        item = self.file_list_widget.itemAt(position)
        if not item: return
        context_menu = QMenu(self)
        context_menu.setStyleSheet("""
            QMenu { background-color: #ffffff; border: 1px solid #cbd5e1; border-radius: 4px; padding: 4px; }
            QMenu::item { padding: 6px 20px; color: #1e293b; font-weight: 500; }
            QMenu::item:selected { background-color: #2563eb; color: #ffffff; border-radius: 3px; }
        """)
        open_action = context_menu.addAction("📂 Open in Excel")
        action = context_menu.exec(self.file_list_widget.mapToGlobal(position))
        if action == open_action:
            self.open_file_externally(item)

    def request_metadata_load(self, item):
        filename = item.text().replace("📊  ", "").strip()
        if filename.startswith("(") or filename.startswith("Reading Error"): return
        folder = self.input_path.text()
        full_file_path = os.path.normpath(os.path.join(folder, filename))
        if not os.path.exists(full_file_path): return

        self.meta_card.setHtml(f"<p style='color: #2563eb; font-weight: bold;'>⏳ Scanning structure metrics for {filename}...</p>")
        self.meta_thread = ExcelMetaWorker(full_file_path)
        self.meta_thread.meta_loaded_signal.connect(self.display_metadata_card)
        self.meta_thread.error_signal.connect(lambda err: self.meta_card.setHtml(f"<p style='color: #ef4444;'>❌ Error: {err}</p>"))
        self.meta_thread.start()

    def display_metadata_card(self, meta):
        html_content = f"""
        <table width="100%" cellpadding="4" style="font-family: 'Segoe UI', Arial; font-size: 13px; color: #1e293b;">
            <tr><td><b>📁 File Size:</b></td><td style="color: #2563eb; font-weight: 600;">{meta['file_size']}</td></tr>
            <tr><td><b>📅 Last Modified:</b></td><td>{meta['mod_date']}</td></tr>
            <tr><td><b>🔢 Total Sheets:</b></td><td>{meta['num_sheets']}</td></tr>
            <tr><td><b>📖 Sheet Names:</b></td><td style="color: #475569; font-style: italic;">{meta['sheet_names']}</td></tr>
            <tr style="background-color: #f8fafc;"><td colspan="2" style="border-top: 1px solid #e2e8f0; padding-top: 6px;"><b>📊 Active Sheet Diagnostics ({meta['active_sheet']}):</b></td></tr>
            <tr><td>&nbsp;&nbsp;&nbsp;&nbsp;• Row Count:</td><td style="font-weight: bold; color: #10b981;">{meta['row_count']} rows</td></tr>
            <tr><td>&nbsp;&nbsp;&nbsp;&nbsp;• Column Count:</td><td style="font-weight: bold; color: #f59e0b;">{meta['col_count']} columns</td></tr>
        </table>
        """
        self.meta_card.setHtml(html_content)

    def populate_file_list(self):
        self.file_list_widget.clear()
        path = self.input_path.text()
        if os.path.exists(path) and os.path.isdir(path):
            try:
                files = [f for f in os.listdir(path) if f.lower().endswith('.xlsx')]
                if files:
                    for filename in files: self.file_list_widget.addItem(f"📊  {filename}")
                else:
                    self.file_list_widget.addItem("(No .xlsx spreadsheets found in path)")
            except Exception as e: self.file_list_widget.addItem(f"Reading Error: {str(e)}")

    def toggle_task_inputs(self):
        task_idx = self.task_selector.currentIndex()

        # Input Visibility Flag Matrix
        is_splitter = task_idx == 0
        is_analysis = task_idx == 4
        needs_template = task_idx in [2, 3]
        needs_equipment = task_idx in [0, 2, 3, 4] 

        self.equipment.setVisible(needs_equipment)
        self.equipment_row_label.setVisible(needs_equipment)
        self.chunk_size.setVisible(is_splitter)
        self.chunk_row_label.setVisible(is_splitter)
        self.template_widget.setVisible(needs_template)
        self.template_row_label.setVisible(needs_template)
        
        # Toggle Particular and Activity inputs visibility
        self.particular_field.setVisible(is_analysis)
        self.particular_row_label.setVisible(is_analysis)
        self.activity_field.setVisible(is_analysis)
        self.activity_row_label.setVisible(is_analysis)

    def select_input_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Source Input Directory")
        if folder:
            self.input_path.setText(folder)
            task_idx = self.task_selector.currentIndex()
            if task_idx == 0: self.output_path.setText(os.path.join(folder, "split_data_files"))
            elif task_idx == 1: self.output_path.setText(os.path.join(folder, "renamed_files"))
            elif task_idx == 2: self.output_path.setText(os.path.join(folder, "collected_data"))
            elif task_idx == 3: self.output_path.setText(os.path.join(folder, "cleaned_data"))
            elif task_idx == 4: self.output_path.setText(os.path.join(folder, "output"))

    def select_output_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Output Target Directory")
        if folder: self.output_path.setText(folder)

    def select_template_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Base Template Excel File", "", "Excel Files (*.xlsx)")
        if file_path: self.template_path_field.setText(file_path)

    def log_message(self, msg):
        self.log.append(msg)

    def execute_current_task(self):
        task_idx = self.task_selector.currentIndex()
        input_folder = self.input_path.text()
        output_folder = self.output_path.text()
        self.progress.setValue(0)
        
        if not input_folder or not output_folder: 
            self.log_message("⚠️ Error: Check input and output paths.")
            return

        # Core Task Router Routing Setup
        if task_idx == 0:
            chunk = self.chunk_size.text()
            try: chunk = int(chunk)
            except ValueError: return
            from splitter_engine import EquipmentProductivity
            self.engine = EquipmentProductivity(folder_path=input_folder, output_folder=output_folder, equipment=self.equipment.currentText().lower(), logger=self.log_message)
            self.engine.chunk_size = chunk
        elif task_idx == 1:
            from naming_engine import FileRenamingEngine
            self.engine = FileRenamingEngine(input_folder=input_folder, file_type=".xlsx")
        elif task_idx == 2:
            template_path = self.template_path_field.text()
            if not template_path or not os.path.exists(template_path): return
            from processing_engine import DataProcessingEngine
            self.engine = DataProcessingEngine(input_folder=input_folder, template_path=template_path, equipment=self.equipment.currentText().lower())
        elif task_idx == 3:
            template_path = self.template_path_field.text()
            if not template_path or not os.path.exists(template_path): return
            from cleaning_engine import DataCleaningEngine
            self.engine = DataCleaningEngine({'input_folder':input_folder, 'output_folder':output_folder, 'template_path':template_path, 'logger':self.log_message, 'progress_callback':self.progress.setValue})
        
        # Route to AnalysisEngine
        elif task_idx == 4:
            equip_name = self.equipment.currentText()
            # Capture dynamic textual inputs directly from the newly provided fields
            user_particular = self.particular_field.text().strip()
            user_activity = self.activity_field.text().strip()

            self.engine = AnalysisEngine(
                data_folder=input_folder,
                template_path=None,
                equipment=equip_name,
                particular=user_particular if user_particular else f"Site Clearing using Dozer {equip_name}",
                activity=user_activity if user_activity else f"Site clearing using {equip_name}"
            )

        # Spawn asynchronous thread handling execution cleanly without freezing UI layout
        self.worker = Worker(self.engine)
        self.worker.log_signal.connect(self.log_message)
        self.worker.progress_signal.connect(self.progress.setValue)
        self.worker.finished_signal.connect(lambda: [self.log_message("✅ Task Processing Complete!"), self.populate_file_list()])
        self.worker.start()

    def open_output_folder(self):
        folder = self.output_path.text()
        if not folder or not os.path.exists(folder): return
        if os.name == "nt": os.startfile(folder)
        elif sys.platform == "darwin": subprocess.run(["open", folder])
        else: subprocess.run(["xdg-open", folder])

    def apply_ui_styles(self):
        self.setStyleSheet("""
            QWidget { font-size: 13px; background-color: #f1f5f9; color: #334155; }
            QLineEdit, QComboBox { border: 1px solid #cbd5e1; border-radius: 6px; padding: 6px; background: white; font-size: 14px; color: #1e293b; }
            QLineEdit:focus, QComboBox:focus, QListWidget:focus { border: 2px solid #2563eb; }
            QTextEdit { border: 1px solid #cbd5e1; border-radius: 6px; background: #1e293b; color: #38bdf8; font-family: 'Consolas', monospace; }
            QPushButton { background-color: #2563eb; color: white; border: none; border-radius: 6px; padding: 8px 14px; font-weight: 600; min-height: 32px; }
            QPushButton:hover { background-color: #1d4ed8; }
            QLabel { font-size: 14px; font-weight: bold; color: #334155; margin-top: 4px; }
            QProgressBar { border: 1px solid #cbd5e1; border-radius: 4px; text-align: center; background: #e2e8f0; font-weight: bold; }
            QProgressBar::chunk { background-color: #2563eb; }
            QListWidget#styledFileList { background-color: white; border: 1px solid #cbd5e1; border-radius: 8px; padding: 5px; outline: none; }
            QListWidget#styledFileList::item { padding: 10px 12px; margin: 2px 4px; border-radius: 6px; background-color: #f8fafc; border: 1px solid #e2e8f0; color: #1e293b; font-weight: 500; font-family: 'Segoe UI', Arial; }
            QListWidget#styledFileList::item:hover { background-color: #eff6ff; color: #2563eb; border: 1px solid #bfdbfe; }
            QListWidget#styledFileList::item:selected { background-color: #2563eb; color: white; font-weight: bold; border: 1px solid #2563eb; }
            QTextEdit#metaSummaryCard { background-color: #ffffff; border: 1px solid #cbd5e1; border-radius: 8px; padding: 8px; }
        """)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = CMIDataProcessorGUI()
    window.show()
    sys.exit(app.exec())